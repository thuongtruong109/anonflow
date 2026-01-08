import sys, asyncio, os
from pathlib import Path
from datetime import datetime
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTextEdit, QLabel, QGroupBox, QGridLayout, QMessageBox,
    QFileDialog, QSplitter, QFrame, QLineEdit, QSpinBox, QCheckBox,
    QScrollArea, QRadioButton, QSizePolicy, QTabWidget, QListWidget, QListWidgetItem
)
from PySide6.QtCore import Qt, QThread, Signal, Slot, QPropertyAnimation, QEasingCurve, QTimer
from PySide6.QtGui import QFont, QTextCursor

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import config
from excel import update_excel_column_a_with_cookie_files, read_excel
from services import create_profile, start_profile, close_profile, delete_profile, get_profiles_list
from runner import run_all_playwright
from utils import safe_print
from handler import process_row

class OutputRedirector:
    """Redirect stdout/stderr to GUI terminal"""
    def __init__(self, text_widget, original_stream):
        self.text_widget = text_widget
        self.original_stream = original_stream

    def write(self, text):
        if text.strip():
            # Also write to original stream for debugging
            if self.original_stream:
                self.original_stream.write(text)

            # Write to GUI
            cursor = self.text_widget.textCursor()
            cursor.movePosition(QTextCursor.End)
            self.text_widget.setTextCursor(cursor)
            self.text_widget.insertPlainText(text)
            self.text_widget.ensureCursorVisible()

    def flush(self):
        if self.original_stream:
            self.original_stream.flush()


class TaskWorker(QThread):
    """Worker thread for running async tasks"""
    log_signal = Signal(str)
    finished_signal = Signal(bool, str)

    def __init__(self, task_name, task_func):
        super().__init__()
        self.task_name = task_name
        self.task_func = task_func
        self._is_stopped = False

    def stop(self):
        """Request the worker to stop"""
        self._is_stopped = True

    def is_stopped(self):
        """Check if stop was requested"""
        return self._is_stopped

    def run(self):
        try:
            self.log_signal.emit(f"\n{'='*70}\n🚀 Starting: {self.task_name}\n{'='*70}\n")

            # Run the task
            self.task_func()

            if self._is_stopped:
                self.log_signal.emit(f"\n{'='*70}\n⏹️ Stopped: {self.task_name}\n{'='*70}\n")
                self.finished_signal.emit(False, f"Task '{self.task_name}' was stopped by user")
            else:
                self.log_signal.emit(f"\n{'='*70}\n✅ Completed: {self.task_name}\n{'='*70}\n")
                self.finished_signal.emit(True, f"Task '{self.task_name}' completed!")

        except Exception as e:
            error_msg = f"Error: {str(e)}"
            self.log_signal.emit(f"\n{'='*70}\n❌ Failed: {self.task_name}\n{error_msg}\n{'='*70}\n")
            self.finished_signal.emit(False, error_msg)


class ModernButton(QPushButton):
    """Styled button with modern look"""
    def __init__(self, text, icon="", color="#2196F3"):
        super().__init__()
        self.base_color = color
        display_text = f"{icon}  {text}" if icon else text
        self.setText(display_text)
        self.setMinimumHeight(35)
        self.setFont(QFont("Segoe UI", 9, QFont.Bold))
        self.setCursor(Qt.PointingHandCursor)
        self.update_style()

    def update_style(self):
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {self.base_color};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 6px 10px;
                text-align: left;
            }}
            QPushButton:hover {{
                background-color: {self.lighten_color(self.base_color, 20)};
            }}
            QPushButton:pressed {{
                background-color: {self.lighten_color(self.base_color, -20)};
            }}
            QPushButton:disabled {{
                background-color: #424242;
                color: #757575;
            }}
        """)

    def lighten_color(self, hex_color, amount):
        hex_color = hex_color.lstrip('#')
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        r = max(0, min(255, r + amount))
        g = max(0, min(255, g + amount))
        b = max(0, min(255, b + amount))
        return f"#{r:02x}{g:02x}{b:02x}"

    def darken_color(self, hex_color, amount=30):
        hex_color = hex_color.lstrip('#')
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        r = max(0, min(255, r - amount))
        g = max(0, min(255, g - amount))
        b = max(0, min(255, b - amount))
        return f"#{r:02x}{g:02x}{b:02x}"


class ProfileListItem(QWidget):
    """Custom widget for profile list items with name and delete button"""
    delete_clicked = Signal(str, str)  # profile_id, profile_name

    def __init__(self, profile_name, profile_id):
        super().__init__()
        self.profile_name = profile_name
        self.profile_id = profile_id
        self.setMinimumHeight(40)  # Ensure adequate height for the widget

        layout = QHBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(10)

        # Profile name label
        self.name_label = QLabel(profile_name)
        self.name_label.setFont(QFont("Segoe UI", 9))
        self.name_label.setStyleSheet("font-weight: bold; background-color: transparent;")

        # Delete button
        self.delete_btn = ModernButton("Delete", "🗑️", "#f44336")
        self.delete_btn.setFixedSize(70, 25)
        self.delete_btn.clicked.connect(self.on_delete_clicked)

        layout.addWidget(self.name_label)
        layout.addStretch()
        layout.addWidget(self.delete_btn)

    def on_delete_clicked(self):
        self.delete_clicked.emit(self.profile_id, self.profile_name)


class GPMMainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.current_worker = None
        self.task_buttons = []
        self.task_checkboxes = []
        self.task_queue = []
        self.task_options = {}
        self.stop_requested = False
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Anonflow")
        self.setMinimumSize(1200, 400)

        self.apply_dark_theme()

        # Main widget
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)

        # Top row: Header + Settings (left) and Tasks (right) with 3:7 ratio
        top_row_layout = QHBoxLayout()
        top_row_layout.setSpacing(6)

        # Left column (30%): Header + Cookie Setting + Tasks
        left_column = QVBoxLayout()
        left_column.setSpacing(6)
        left_column.addWidget(self.create_header())
        left_column.addWidget(self.create_cookie_setting_section())
        left_column.addWidget(self.create_tasks_section())

        # Create a widget to hold the left column
        left_widget = QWidget()
        left_widget.setLayout(left_column)
        left_widget.setMinimumWidth(270)  # Fixed minimum width for quick settings, cookie setting and tasks
        left_widget.setMaximumWidth(270)  # Maximum width limit

        # Add to top row - proxy settings and behavior actions will take remaining space
        top_row_layout.addWidget(left_widget)
        proxy_widget = self.create_config_section()
        top_row_layout.addWidget(proxy_widget)
        behavior_widget = self.create_behavior_actions_section()
        top_row_layout.addWidget(behavior_widget)
        layout.addLayout(top_row_layout)

        # Main splitter - Terminal only
        splitter = QSplitter(Qt.Vertical)

        # Bottom section: Terminal
        splitter.addWidget(self.create_terminal_section())
        layout.addWidget(splitter)

    def apply_dark_theme(self):
        self.setStyleSheet("""
            QMainWindow, QWidget {
                background-color: #0d0d0d;
                color: #e0e0e0;
                font-family: 'Segoe UI', Arial;
            }
            QGroupBox {
                border: 2px solid #2196F3;
                border-radius: 8px;
                margin-top: 6px;
                padding-top: 6px;
                font-weight: bold;
                font-size: 10pt;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                color: #2196F3;
            }
            QTextEdit {
                background-color: #000000;
                color: #00ff00;
                border: 2px solid #2196F3;
                border-radius: 6px;
                padding: 6px;
                font-family: 'Consolas', 'Courier New', monospace;
                font-size: 9pt;
                selection-background-color: #2196F3;
            }
            QLineEdit {
                background-color: #1a1a1a;
                color: #ffffff;
                border: 1px solid #333;
                border-radius: 4px;
                padding: 6px;
                font-size: 9pt;
            }
            QLineEdit:focus {
                border: 1px solid #2196F3;
            }
            QSpinBox {
                background-color: #1a1a1a;
                color: #ffffff;
                border: 1px solid #333;
                border-radius: 4px;
                padding: 4px;
                font-size: 9pt;
            }
            QSpinBox:focus {
                border: 1px solid #2196F3;
            }
            QCheckBox {
                spacing: 5px;
                color: #e0e0e0;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
                border-radius: 3px;
                border: 1px solid #2196F3;
            }
            QCheckBox::indicator:checked {
                background-color: #2196F3;
            }
            QLabel {
                color: #e0e0e0;
            }
            QSplitter::handle {
                background-color: transparent;
                height: 0px;
            }
            QScrollBar:vertical {
                background-color: #1a1a1a;
                width: 8px;
                border-radius: 1rem;
            }
            QScrollBar::handle:vertical {
                background-color: #555555;
                border-radius: 1rem;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #777777;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                background: none;
                border: none;
            }
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                background: none;
            }
        """)

    def create_header(self):
        frame = QFrame()
        frame.setMaximumHeight(70)
        frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #1a237e, stop:1 #2196F3);
                border-radius: 8px;
                padding: 2px 4px;
            }
        """)
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        title = QLabel("🚀 GPM Automation")
        title.setFont(QFont("Segoe UI", 12, QFont.Bold))
        title.setStyleSheet("color: white;")

        subtitle = QLabel("TikTok Profile & Cookie Management")
        subtitle.setFont(QFont("Segoe UI", 8))
        subtitle.setStyleSheet("color: #bbdefb;")

        layout.addWidget(title)
        layout.addWidget(subtitle)
        return frame

    def create_config_section(self):
        group = QGroupBox("⚙️ Profile Setting")
        group.setStyleSheet("""
            QGroupBox {
                background-color: #1a1a1a;
                border: 2px solid #2196F3;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                font-weight: bold;
                font-size: 10pt;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                color: #2196F3;
            }
        """)
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(8, 8, 8, 8)

        # Create tab widget
        tab_widget = QTabWidget()
        tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: none;
                background-color: #1a1a1a;
            }
            QTabBar::tab {
                background-color: #2a2a2a;
                color: #ffffff;
                padding: 8px 16px;
                border: 1px solid #2196F3;
                border-bottom: none;
                border-radius: 4px 4px 0 0;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background-color: #2196F3;
                color: #ffffff;
            }
            QTabBar::tab:hover {
                background-color: #42A5F5;
            }
        """)

        # Proxies tab - Management buttons and text area
        proxies_tab = QWidget()
        proxies_layout = QVBoxLayout(proxies_tab)
        proxies_layout.setSpacing(10)
        proxies_layout.setContentsMargins(8, 8, 8, 8)

        # Profile addresses text area
        self.profiles_text = QTextEdit()
        self.profiles_text.setReadOnly(False)
        self.profiles_text.setFont(QFont("Consolas", 8))
        self.profiles_text.setStyleSheet("""
            QTextEdit {
                background-color: #0a0a0a;
                color: #ffaa00;
                border: none;
                padding: 6px;
                font-family: 'Consolas', 'Courier New', monospace;
            }
            QTextEdit:focus {
                border: 1px solid #42A5F5;
            }
        """)
        self.profiles_text.setPlaceholderText("Profile addresses will be displayed here...\nEdit and save to update Excel file.")

        proxies_layout.addWidget(self.profiles_text)

        # Action buttons (2x2 grid)
        buttons_layout = QGridLayout()
        buttons_layout.setSpacing(8)

        create_profiles_btn = ModernButton("Create Profiles", "➕", "#2196F3")
        create_profiles_btn.clicked.connect(self.create_profiles)

        update_profiles_btn = ModernButton("Update Profiles", "🔄", "#00BCD4")
        update_profiles_btn.clicked.connect(self.update_profiles)

        view_btn = ModernButton("View File", "📊", "#9C27B0")
        view_btn.clicked.connect(self.view_excel_file)

        save_btn = ModernButton("Save Changes", "💾", "#4CAF50")
        save_btn.clicked.connect(self.save_profiles_to_excel)

        buttons_layout.addWidget(create_profiles_btn, 0, 0)
        buttons_layout.addWidget(update_profiles_btn, 0, 1)
        buttons_layout.addWidget(view_btn, 1, 0)
        buttons_layout.addWidget(save_btn, 1, 1)

        proxies_layout.addLayout(buttons_layout)

        tab_widget.addTab(proxies_tab, "Proxies")

        # Profiles tab - Display list of existing profiles
        profiles_tab = QWidget()
        profiles_layout = QVBoxLayout(profiles_tab)
        profiles_layout.setSpacing(10)
        profiles_layout.setContentsMargins(8, 8, 8, 8)

        # Profiles list widget
        self.profiles_list = QListWidget()
        self.profiles_list.setStyleSheet("""
            QListWidget {
                background-color: #0a0a0a;
                color: #ffaa00;
                border: none;
                padding: 6px;
                font-family: 'Consolas', 'Courier New', monospace;
                font-size: 9pt;
            }
            QListWidget::item {
                padding: 0 8px 8px 8px;
                border-bottom: 1px solid #333;
            }
            QListWidget::item:selected {
                background-color: #222222;
                color: white;
            }
            QScrollBar:vertical {
                border: none;
                background-color: #1a1a1a;
                width: 8px;
                border-radius: 1rem;
            }
            QScrollBar::handle:vertical {
                background-color: #555555;
                border-radius: 1rem;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #777777;
            }
            QScrollBar:horizontal {
                border: none;
                background-color: #1a1a1a;
                height: 8px;
                border-radius: 1rem;
            }
            QScrollBar::handle:horizontal {
                background-color: #555555;
                border-radius: 1rem;
                min-width: 20px;
            }
            QScrollBar::handle:horizontal:hover {
                background-color: #777777;
            }
        """)

        profiles_layout.addWidget(self.profiles_list)

        # Refresh button for profiles list
        refresh_profiles_btn = ModernButton("Refresh Profiles", "🔄", "#00BCD4")
        refresh_profiles_btn.clicked.connect(self.refresh_profiles_list)

        profiles_layout.addWidget(refresh_profiles_btn)

        tab_widget.addTab(profiles_tab, "Profiles")

        # Others tab - Configuration options
        others_tab = QWidget()
        others_layout = QVBoxLayout(others_tab)
        others_layout.setSpacing(10)
        others_layout.setContentsMargins(8, 8, 8, 8)

        # Threads and Start Limit on same row
        settings_layout = QHBoxLayout()
        settings_layout.setSpacing(10)

        # Threads
        threads_label = QLabel("Threads:")
        threads_label.setFont(QFont("Segoe UI", 9))
        self.threads_spin = QSpinBox()
        self.threads_spin.setRange(1, 20)
        self.threads_spin.setValue(config.THREADS)
        self.threads_spin.valueChanged.connect(self.auto_save_threads)
        self.threads_spin.setMaximumWidth(80)

        settings_layout.addWidget(threads_label)
        settings_layout.addWidget(self.threads_spin)

        # Start Limit
        limit_label = QLabel("Start Limit:")
        limit_label.setFont(QFont("Segoe UI", 9))

        self.start_limit = QSpinBox()
        self.start_limit.setRange(1, 50)
        self.start_limit.setValue(config.START_LIMIT)
        self.start_limit.valueChanged.connect(self.auto_save_start_limit)
        self.start_limit.setMaximumWidth(80)

        settings_layout.addWidget(limit_label)
        settings_layout.addWidget(self.start_limit)
        settings_layout.addStretch()

        others_layout.addLayout(settings_layout)

        # Run with CDP checkbox
        self.run_with_cdp_checkbox = QCheckBox("Run with CDP")
        self.run_with_cdp_checkbox.setChecked(True)  # Default is checked (True)
        self.run_with_cdp_checkbox.setFont(QFont("Segoe UI", 9))
        self.run_with_cdp_checkbox.setStyleSheet("""
            QCheckBox {
                color: #ffffff;
                spacing: 5px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #2196F3;
                background-color: #1a1a1a;
                border-radius: 3px;
            }
            QCheckBox::indicator:checked {
                border: 2px solid #2196F3;
                background-color: #2196F3;
                border-radius: 3px;
            }
        """)
        others_layout.addWidget(self.run_with_cdp_checkbox)

        # Full screen size checkbox
        self.full_screen_checkbox = QCheckBox("Full screen size")
        self.full_screen_checkbox.setChecked(config.FULL_SCREEN)
        self.full_screen_checkbox.setFont(QFont("Segoe UI", 9))
        self.full_screen_checkbox.stateChanged.connect(self.auto_save_full_screen)
        self.full_screen_checkbox.setStyleSheet("""
            QCheckBox {
                color: #ffffff;
                spacing: 5px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #2196F3;
                background-color: #1a1a1a;
                border-radius: 3px;
            }
            QCheckBox::indicator:checked {
                border: 2px solid #2196F3;
                background-color: #2196F3;
                border-radius: 3px;
            }
        """)
        others_layout.addWidget(self.full_screen_checkbox)

        # Add stretch to push content to top
        others_layout.addStretch()

        tab_widget.addTab(others_tab, "Others")

        main_layout.addWidget(tab_widget)

        group.setLayout(main_layout)

        self.load_excel_data()
        self.refresh_profiles_list()  # Load profiles list on startup

        return group

    def create_tasks_section(self):
        group = QGroupBox("📋 Tasks")
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                border: none;
                background-color: #1a1a1a;
                width: 8px;
                border-radius: 1rem;
            }
            QScrollBar::handle:vertical {
                background-color: #555555;
                border-radius: 1rem;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #777777;
            }
        """)

        # Container for all tasks
        tasks_container = QWidget()
        tasks_layout = QVBoxLayout(tasks_container)
        tasks_layout.setSpacing(6)

        # Define tasks with their metadata (name, icon, original_index, has_options, color)
        tasks = [
            ("Start Profiles", "▶️", 0, False, "#4CAF50"),  # Green - single button
            ("Close Profiles", "⏹️", 1, False, "#FF9800"),         # Orange
            ("Delete Profiles", "🗑️", 2, False, "#F44336"),       # Red
        ]

        self.task_buttons = []
        self.task_options = {}  # Store option widgets for each task

        # Create and add stop button first (will be positioned next to Start Profiles)
        self.stop_btn = ModernButton("🛑  Stop Tasks", "", "#F44336")
        self.stop_btn.setMinimumHeight(32)
        self.stop_btn.setFont(QFont("Segoe UI", 9))
        self.stop_btn.clicked.connect(self.stop_tasks)
        self.stop_btn.setEnabled(False)  # Disabled by default

        for i, (name, icon, idx, has_options, color) in enumerate(tasks):
            # Create task item container (remove wrapper, just use button directly)
            button = ModernButton(f"{icon}  {name}", "", color)
            button.setMinimumHeight(32)
            button.setFont(QFont("Segoe UI", 9))

            # Add arrow button if task has options
            arrow_btn = None
            options_widget = None
            if has_options:
                # Create a container for button and arrow
                task_container = QWidget()
                task_layout = QHBoxLayout(task_container)
                task_layout.setContentsMargins(0, 0, 0, 0)
                task_layout.setSpacing(4)

                task_layout.addWidget(button)

                arrow_btn = QPushButton("▼")
                arrow_btn.setMaximumWidth(30)
                arrow_btn.setMaximumHeight(32)
                arrow_btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {color};
                        color: white;
                        border: none;
                        border-radius: 4px;
                        font-size: 12px;
                        font-weight: bold;
                    }}
                    QPushButton:hover {{
                        background-color: {self.lighten_color(color)};
                    }}
                    QPushButton:pressed {{
                        background-color: {self.darken_color(color)};
                    }}
                """)
                arrow_btn.setCursor(Qt.PointingHandCursor)
                task_layout.addWidget(arrow_btn)

                # Create options widget (initially hidden)
                options_widget = QWidget()
                options_widget.setVisible(False)
                options_widget.setStyleSheet("""
                    QWidget {
                        border: none;
                        background-color: transparent;
                    }
                """)
                options_layout = QVBoxLayout(options_widget)
                options_layout.setContentsMargins(20, 4, 4, 4)
                options_layout.setSpacing(4)

                # Add specific options based on task
                if name == "Run Automation":
                    # Add follower input as example
                    follower_layout = QHBoxLayout()
                    follower_label = QLabel("👥 Target Followers:")
                    follower_label.setFont(QFont("Segoe UI", 9))
                    follower_label.setStyleSheet("color: #bbdefb;")

                    follower_input = QSpinBox()
                    follower_input.setRange(1, 10000)
                    follower_input.setValue(100)
                    follower_input.setStyleSheet("""
                        QSpinBox {
                            background-color: #0d0d0d;
                            color: #2196F3;
                            border: 1px solid #2196F3;
                            border-radius: 4px;
                            padding: 4px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QSpinBox:focus {
                            border: 2px solid #42A5F5;
                        }
                    """)
                    follower_input.setMinimumWidth(100)

                    follower_layout.addWidget(follower_label)
                    follower_layout.addWidget(follower_input)
                    follower_layout.addStretch()

                    options_layout.addLayout(follower_layout)

                    # Store reference to follower input
                    self.task_options[name] = {
                        'widget': options_widget,
                        'follower_input': follower_input
                    }

                elif name == "Follow":
                    # Follow mode selection
                    mode_layout = QVBoxLayout()
                    mode_layout.setSpacing(4)

                    # Radio buttons for follow modes
                    radio_layout = QHBoxLayout()
                    radio_layout.setSpacing(10)

                    random_radio = QRadioButton("Random")
                    random_radio.setFont(QFont("Segoe UI", 9))
                    random_radio.setStyleSheet("""
                        QRadioButton {
                            color: #e0e0e0;
                            spacing: 6px;
                        }
                        QRadioButton::indicator {
                            width: 12px;
                            height: 12px;
                            border-radius: 8px;
                            border: 2px solid #2196F3;
                            background-color: #1a1a1a;
                        }
                        QRadioButton::indicator:checked {
                            background-color: #2196F3;
                            border: 2px solid #2196F3;
                        }
                        QRadioButton::indicator:hover {
                            border: 2px solid #42A5F5;
                        }
                    """)

                    mutual_radio = QRadioButton("Mutual")
                    mutual_radio.setFont(QFont("Segoe UI", 9))
                    mutual_radio.setStyleSheet("""
                        QRadioButton {
                            color: #e0e0e0;
                            spacing: 6px;
                        }
                        QRadioButton::indicator {
                            width: 12px;
                            height: 12px;
                            border-radius: 8px;
                            border: 2px solid #2196F3;
                            background-color: #1a1a1a;
                        }
                        QRadioButton::indicator:checked {
                            background-color: #2196F3;
                            border: 2px solid #2196F3;
                        }
                        QRadioButton::indicator:hover {
                            border: 2px solid #42A5F5;
                        }
                    """)

                    target_radio = QRadioButton("Target")
                    target_radio.setFont(QFont("Segoe UI", 9))
                    target_radio.setStyleSheet("""
                        QRadioButton {
                            color: #e0e0e0;
                            spacing: 6px;
                        }
                        QRadioButton::indicator {
                            width: 12px;
                            height: 12px;
                            border-radius: 8px;
                            border: 2px solid #2196F3;
                            background-color: #1a1a1a;
                        }
                        QRadioButton::indicator:checked {
                            background-color: #2196F3;
                            border: 2px solid #2196F3;
                        }
                        QRadioButton::indicator:hover {
                            border: 2px solid #42A5F5;
                        }
                    """)

                    # Set random as default
                    random_radio.setChecked(True)

                    radio_layout.addWidget(random_radio)
                    radio_layout.addWidget(mutual_radio)
                    radio_layout.addWidget(target_radio)
                    radio_layout.addStretch()

                    mode_layout.addLayout(radio_layout)

                    # Target username input (initially hidden)
                    target_layout = QHBoxLayout()

                    target_input = QLineEdit()
                    target_input.setPlaceholderText("Enter @username...")
                    target_input.setStyleSheet("""
                        QLineEdit {
                            background-color: #0d0d0d;
                            color: #2196F3;
                            border: 1px solid #2196F3;
                            border-radius: 4px;
                            padding: 4px;
                            font-size: 9pt;
                            margin-top: 4px;
                        }
                    """)
                    target_input.setMinimumWidth(150)

                    target_layout.addWidget(target_input)
                    target_layout.addStretch()

                    # Initially hide target input
                    target_input.setVisible(False)

                    mode_layout.addLayout(target_layout)
                    options_layout.addLayout(mode_layout)

                    # Connect radio button signals to show/hide target input
                    def toggle_target_input():
                        is_target = target_radio.isChecked()
                        target_input.setVisible(is_target)

                    random_radio.toggled.connect(toggle_target_input)
                    mutual_radio.toggled.connect(toggle_target_input)
                    target_radio.toggled.connect(toggle_target_input)

                    # Store references
                    self.task_options[name] = {
                        'widget': options_widget,
                        'random_radio': random_radio,
                        'mutual_radio': mutual_radio,
                        'target_radio': target_radio,
                        'target_input': target_input
                    }

                elif name == "View video":
                    # View video options
                    view_layout = QVBoxLayout()
                    view_layout.setSpacing(4)

                    # Amount loop input
                    amount_layout = QHBoxLayout()
                    amount_label = QLabel("👀 Amount loop:")
                    amount_label.setFont(QFont("Segoe UI", 9))
                    amount_label.setStyleSheet("color: #bbdefb;")

                    amount_input = QSpinBox()
                    amount_input.setRange(1, 50)
                    amount_input.setValue(5)
                    amount_input.setStyleSheet("""
                        QSpinBox {
                            background-color: #0d0d0d;
                            color: #2196F3;
                            border: 1px solid #2196F3;
                            border-radius: 4px;
                            padding: 4px;
                            font-size: 9pt;
                        }
                        QSpinBox:focus {
                            border: 2px solid #42A5F5;
                        }
                    """)
                    amount_input.setMinimumWidth(80)

                    amount_layout.addWidget(amount_label)
                    amount_layout.addWidget(amount_input)
                    amount_layout.addStretch()

                    view_layout.addLayout(amount_layout)

                    # Radio buttons for view modes
                    radio_layout = QHBoxLayout()
                    radio_layout.setSpacing(10)

                    random_radio = QRadioButton("Random")
                    random_radio.setFont(QFont("Segoe UI", 9))
                    random_radio.setStyleSheet("""
                        QRadioButton {
                            color: #e0e0e0;
                            spacing: 6px;
                        }
                        QRadioButton::indicator {
                            width: 12px;
                            height: 12px;
                            border-radius: 8px;
                            border: 2px solid #2196F3;
                            background-color: #1a1a1a;
                        }
                        QRadioButton::indicator:checked {
                            background-color: #2196F3;
                            border: 2px solid #2196F3;
                        }
                        QRadioButton::indicator:hover {
                            border: 2px solid #42A5F5;
                        }
                    """)

                    mutual_radio = QRadioButton("Mutual")
                    mutual_radio.setFont(QFont("Segoe UI", 9))
                    mutual_radio.setStyleSheet("""
                        QRadioButton {
                            color: #e0e0e0;
                            spacing: 6px;
                        }
                        QRadioButton::indicator {
                            width: 12px;
                            height: 12px;
                            border-radius: 8px;
                            border: 2px solid #2196F3;
                            background-color: #1a1a1a;
                        }
                        QRadioButton::indicator:checked {
                            background-color: #2196F3;
                            border: 2px solid #2196F3;
                        }
                        QRadioButton::indicator:hover {
                            border: 2px solid #42A5F5;
                        }
                    """)

                    # Set random as default
                    random_radio.setChecked(True)

                    radio_layout.addWidget(random_radio)
                    radio_layout.addWidget(mutual_radio)
                    radio_layout.addStretch()

                    view_layout.addLayout(radio_layout)
                    options_layout.addLayout(view_layout)

                    # Store references
                    self.task_options[name] = {
                        'widget': options_widget,
                        'amount_input': amount_input,
                        'random_radio': random_radio,
                        'mutual_radio': mutual_radio
                    }

                # Connect arrow button to toggle options
                def make_toggle_func(arrow, opts, task_name):
                    def toggle():
                        is_visible = opts.isVisible()
                        opts.setVisible(not is_visible)
                        arrow.setText("▲" if not is_visible else "▼")
                    return toggle

                arrow_btn.clicked.connect(make_toggle_func(arrow_btn, options_widget, name))

                # Create a vertical container for button+options
                task_item = QWidget()
                task_item_layout = QVBoxLayout(task_item)
                task_item_layout.setContentsMargins(0, 0, 0, 0)
                task_item_layout.setSpacing(4)

                task_item_layout.addWidget(task_container)

                # Add options widget if exists
                if options_widget:
                    task_item_layout.addWidget(options_widget)

            else:
                # No options, just use the button directly
                task_item = button

            # Add task to vertical layout
            tasks_layout.addWidget(task_item)

            # Add stop button after Start Profiles
            if i == 0:  # Start Profiles
                tasks_layout.addWidget(self.stop_btn)

            self.task_buttons.append({
                'button': button,
                'name': name,
                'icon': icon,
                'index': idx
            })

            # Connect button to run task
            button.clicked.connect(lambda checked, n=name, i=icon, x=idx: self.run_single_task(n, i, x))

        # Remove the stretch since we're using grid layout
        scroll.setWidget(tasks_container)
        layout.addWidget(scroll)

        group.setLayout(layout)
        return group

    def create_behavior_actions_section(self):
        group = QGroupBox("🎭 Behavior Actions")
        layout = QVBoxLayout()
        layout.setSpacing(8)
        layout.setContentsMargins(6, 6, 6, 6)

        # Mode selection radio buttons
        mode_layout = QHBoxLayout()
        mode_layout.setSpacing(10)
        mode_layout.setContentsMargins(0, 0, 0, 0)

        mode_label = QLabel("Mode:")
        mode_label.setFont(QFont("Segoe UI", 10, QFont.Bold))
        mode_label.setStyleSheet("color: #e0e0e0;")

        self.tiktok_radio = QRadioButton("Tiktok")
        self.tiktok_radio.setFont(QFont("Segoe UI", 9))
        self.tiktok_radio.setChecked(True)  # Default to Tiktok
        self.tiktok_radio.setStyleSheet("""
            QRadioButton {
                color: #e0e0e0;
                spacing: 6px;
            }
            QRadioButton::indicator {
                width: 14px;
                height: 14px;
                border-radius: 8px;
                border: 2px solid #2196F3;
                background-color: #1a1a1a;
            }
            QRadioButton::indicator:checked {
                background-color: #2196F3;
                border: 2px solid #2196F3;
            }
            QRadioButton::indicator:hover {
                border: 2px solid #42A5F5;
            }
        """)

        self.search_radio = QRadioButton("Search")
        self.search_radio.setFont(QFont("Segoe UI", 9))
        self.search_radio.setStyleSheet("""
            QRadioButton {
                color: #e0e0e0;
                spacing: 6px;
            }
            QRadioButton::indicator {
                width: 14px;
                height: 14px;
                border-radius: 8px;
                border: 2px solid #2196F3;
                background-color: #1a1a1a;
            }
            QRadioButton::indicator:checked {
                background-color: #2196F3;
                border: 2px solid #2196F3;
            }
            QRadioButton::indicator:hover {
                border: 2px solid #42A5F5;
            }
        """)

        mode_layout.addWidget(mode_label)
        mode_layout.addWidget(self.tiktok_radio)
        mode_layout.addWidget(self.search_radio)
        mode_layout.addStretch()

        # Align all items to top to prevent floating in the middle
        mode_layout.setAlignment(mode_label, Qt.AlignTop)
        mode_layout.setAlignment(self.tiktok_radio, Qt.AlignTop)
        mode_layout.setAlignment(self.search_radio, Qt.AlignTop)

        layout.addLayout(mode_layout)

        # Search time input (initially hidden) - giờ và phút
        self.search_time_layout = QHBoxLayout()
        self.search_time_layout.setContentsMargins(20, 0, 0, 10)

        search_time_label = QLabel("⏱️ Search time:")
        search_time_label.setFont(QFont("Segoe UI", 9))
        search_time_label.setStyleSheet("color: #bbdefb;")

        # Ô input cho giờ
        self.search_hours_input = QSpinBox()
        self.search_hours_input.setRange(0, 24)
        self.search_hours_input.setValue(0)
        self.search_hours_input.setStyleSheet("""
            QSpinBox {
                background-color: #0d0d0d;
                color: #2196F3;
                border: 1px solid #2196F3;
                border-radius: 4px;
                padding: 4px;
                font-size: 9pt;
            }
            QSpinBox:focus {
                border: 2px solid #42A5F5;
            }
        """)
        self.search_hours_input.setMinimumWidth(60)

        hours_label = QLabel("hours")
        hours_label.setFont(QFont("Segoe UI", 8))
        hours_label.setStyleSheet("color: #bbdefb; margin-left: 2px; margin-right: 8px;")

        # Ô input cho phút
        self.search_minutes_input = QSpinBox()
        self.search_minutes_input.setRange(1, 59)
        self.search_minutes_input.setValue(5)
        self.search_minutes_input.setStyleSheet("""
            QSpinBox {
                background-color: #0d0d0d;
                color: #2196F3;
                border: 1px solid #2196F3;
                border-radius: 4px;
                padding: 4px;
                font-size: 9pt;
            }
            QSpinBox:focus {
                border: 2px solid #42A5F5;
            }
        """)
        self.search_minutes_input.setMinimumWidth(60)

        minutes_label = QLabel("minutes")
        minutes_label.setFont(QFont("Segoe UI", 8))
        minutes_label.setStyleSheet("color: #bbdefb; margin-left: 2px;")

        self.search_time_layout.addWidget(search_time_label)
        self.search_time_layout.addWidget(self.search_hours_input)
        self.search_time_layout.addWidget(hours_label)
        self.search_time_layout.addWidget(self.search_minutes_input)
        self.search_time_layout.addWidget(minutes_label)
        self.search_time_layout.addStretch()

        # Initially hide search time input
        for i in range(5):  # 5 widgets: label, hours_input, hours_label, minutes_input, minutes_label
            self.search_time_layout.itemAt(i).widget().setVisible(False)

        # Wrap search time layout in widget for alignment
        search_time_widget = QWidget()
        search_time_widget.setLayout(self.search_time_layout)
        search_time_widget.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Minimum)

        layout.addWidget(search_time_widget, alignment=Qt.AlignTop)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                border: none;
                background-color: #1a1a1a;
                width: 8px;
                border-radius: 1rem;
            }
            QScrollBar::handle:vertical {
                background-color: #555555;
                border-radius: 1rem;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #777777;
            }
        """)

        # Container for all tasks
        tasks_container = QWidget()
        tasks_layout = QVBoxLayout(tasks_container)
        tasks_layout.setSpacing(6)
        tasks_layout.setContentsMargins(0, 0, 10, 0)

        # Define tasks with their metadata (name, icon, original_index, has_options)
        tasks = [
            ("Follow", "👥", 7, True),
            ("View video", "👀", 8, True),
            ("Like video", "❤️", 9, False),
            ("Comment", "💬", 10, False),
        ]

        # Note: We don't reset self.task_checkboxes here, we append to it
        # self.task_options will be updated

        for i, (name, icon, idx, has_options) in enumerate(tasks):
            # Create task item container
            task_item = QFrame()
            task_item.setStyleSheet("""
                QFrame {
                    background-color: #1a1a1a;
                    border: 1px solid #333;
                    border-radius: 6px;
                    padding: 4px;
                }
                QFrame:hover {
                    border: 1px solid #2196F3;
                }
            """)
            task_item_layout = QVBoxLayout(task_item)
            task_item_layout.setContentsMargins(4, 4, 4, 4)
            task_item_layout.setSpacing(4)

            # Top row: Checkbox and arrow button
            top_row = QHBoxLayout()
            top_row.setSpacing(4)

            checkbox = QCheckBox(f"{icon}  {name}")
            checkbox.setFont(QFont("Segoe UI", 9))
            checkbox.setStyleSheet("""
                QCheckBox {
                    spacing: 8px;
                    color: #e0e0e0;
                    padding: 4px;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                    border-radius: 3px;
                    border: 2px solid #2196F3;
                    background-color: #1a1a1a;
                }
                QCheckBox::indicator:checked {
                    background-color: #2196F3;
                    border: 2px solid #2196F3;
                }
                QCheckBox::indicator:hover {
                    border: 2px solid #42A5F5;
                }
                QCheckBox:hover {
                    color: #42A5F5;
                }
            """)

            top_row.addWidget(checkbox)

            # Add arrow button if task has options
            arrow_btn = None
            options_widget = None
            if has_options:
                arrow_btn = QPushButton("▼")
                arrow_btn.setMaximumWidth(30)
                arrow_btn.setMaximumHeight(30)
                arrow_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #555555;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        font-size: 10px;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background-color: #777777;
                    }
                    QPushButton:pressed {
                        background-color: #555555;
                    }
                """)
                arrow_btn.setCursor(Qt.PointingHandCursor)
                top_row.addWidget(arrow_btn)

                # Create options widget (initially hidden)
                options_widget = QWidget()
                options_widget.setVisible(False)
                options_widget.setStyleSheet("""
                    QWidget {
                        border: none;
                        background-color: transparent;
                    }
                """)
                options_layout = QVBoxLayout(options_widget)
                options_layout.setContentsMargins(20, 4, 4, 4)
                options_layout.setSpacing(4)

                # Add specific options based on task
                if name == "Follow":
                    # Follow mode selection
                    mode_layout = QVBoxLayout()
                    mode_layout.setSpacing(4)

                    # Radio buttons for follow modes
                    radio_layout = QHBoxLayout()
                    radio_layout.setSpacing(10)

                    random_radio = QRadioButton("Random")
                    random_radio.setFont(QFont("Segoe UI", 9))
                    random_radio.setStyleSheet("""
                        QRadioButton {
                            color: #e0e0e0;
                            spacing: 6px;
                        }
                        QRadioButton::indicator {
                            width: 12px;
                            height: 12px;
                            border-radius: 8px;
                            border: 2px solid #2196F3;
                            background-color: #1a1a1a;
                        }
                        QRadioButton::indicator:checked {
                            background-color: #2196F3;
                            border: 2px solid #2196F3;
                        }
                        QRadioButton::indicator:hover {
                            border: 2px solid #42A5F5;
                        }
                    """)

                    mutual_radio = QRadioButton("Mutual")
                    mutual_radio.setFont(QFont("Segoe UI", 9))
                    mutual_radio.setStyleSheet("""
                        QRadioButton {
                            color: #e0e0e0;
                            spacing: 6px;
                        }
                        QRadioButton::indicator {
                            width: 12px;
                            height: 12px;
                            border-radius: 8px;
                            border: 2px solid #2196F3;
                            background-color: #1a1a1a;
                        }
                        QRadioButton::indicator:checked {
                            background-color: #2196F3;
                            border: 2px solid #2196F3;
                        }
                        QRadioButton::indicator:hover {
                            border: 2px solid #42A5F5;
                        }
                    """)

                    target_radio = QRadioButton("Target")
                    target_radio.setFont(QFont("Segoe UI", 9))
                    target_radio.setStyleSheet("""
                        QRadioButton {
                            color: #e0e0e0;
                            spacing: 6px;
                        }
                        QRadioButton::indicator {
                            width: 12px;
                            height: 12px;
                            border-radius: 8px;
                            border: 2px solid #2196F3;
                            background-color: #1a1a1a;
                        }
                        QRadioButton::indicator:checked {
                            background-color: #2196F3;
                            border: 2px solid #2196F3;
                        }
                        QRadioButton::indicator:hover {
                            border: 2px solid #42A5F5;
                        }
                    """)

                    # Set random as default
                    random_radio.setChecked(True)

                    radio_layout.addWidget(random_radio)
                    radio_layout.addWidget(mutual_radio)
                    radio_layout.addWidget(target_radio)
                    radio_layout.addStretch()

                    mode_layout.addLayout(radio_layout)

                    # Target username input (initially hidden)
                    target_layout = QHBoxLayout()

                    target_input = QLineEdit()
                    target_input.setPlaceholderText("Enter @username...")
                    target_input.setStyleSheet("""
                        QLineEdit {
                            background-color: #0d0d0d;
                            color: #2196F3;
                            border: 1px solid #2196F3;
                            border-radius: 4px;
                            padding: 4px;
                            font-size: 9pt;
                            margin-top: 4px;
                        }
                    """)
                    target_input.setMinimumWidth(150)

                    target_layout.addWidget(target_input)
                    target_layout.addStretch()

                    # Initially hide target input
                    target_input.setVisible(False)

                    mode_layout.addLayout(target_layout)
                    options_layout.addLayout(mode_layout)

                    # Connect radio button signals to show/hide target input
                    def toggle_target_input():
                        is_target = target_radio.isChecked()
                        target_input.setVisible(is_target)

                    random_radio.toggled.connect(toggle_target_input)
                    mutual_radio.toggled.connect(toggle_target_input)
                    target_radio.toggled.connect(toggle_target_input)

                    # Store references
                    self.task_options[name] = {
                        'widget': options_widget,
                        'random_radio': random_radio,
                        'mutual_radio': mutual_radio,
                        'target_radio': target_radio,
                        'target_input': target_input
                    }

                elif name == "View video":
                    # View video options
                    view_layout = QVBoxLayout()
                    view_layout.setSpacing(4)

                    # Amount loop input
                    amount_layout = QHBoxLayout()
                    amount_label = QLabel("👀 Amount loop:")
                    amount_label.setFont(QFont("Segoe UI", 9))
                    amount_label.setStyleSheet("color: #bbdefb;")

                    amount_input = QSpinBox()
                    amount_input.setRange(1, 50)
                    amount_input.setValue(5)
                    amount_input.setStyleSheet("""
                        QSpinBox {
                            background-color: #0d0d0d;
                            color: #2196F3;
                            border: 1px solid #2196F3;
                            border-radius: 4px;
                            padding: 4px;
                            font-size: 9pt;
                        }
                        QSpinBox:focus {
                            border: 2px solid #42A5F5;
                        }
                    """)
                    amount_input.setMinimumWidth(80)

                    amount_layout.addWidget(amount_label)
                    amount_layout.addWidget(amount_input)
                    amount_layout.addStretch()

                    view_layout.addLayout(amount_layout)

                    # Radio buttons for view modes
                    radio_layout = QHBoxLayout()
                    radio_layout.setSpacing(10)

                    random_radio = QRadioButton("Random")
                    random_radio.setFont(QFont("Segoe UI", 9))
                    random_radio.setStyleSheet("""
                        QRadioButton {
                            color: #e0e0e0;
                            spacing: 6px;
                        }
                        QRadioButton::indicator {
                            width: 12px;
                            height: 12px;
                            border-radius: 8px;
                            border: 2px solid #2196F3;
                            background-color: #1a1a1a;
                        }
                        QRadioButton::indicator:checked {
                            background-color: #2196F3;
                            border: 2px solid #2196F3;
                        }
                        QRadioButton::indicator:hover {
                            border: 2px solid #42A5F5;
                        }
                    """)

                    mutual_radio = QRadioButton("Mutual")
                    mutual_radio.setFont(QFont("Segoe UI", 9))
                    mutual_radio.setStyleSheet("""
                        QRadioButton {
                            color: #e0e0e0;
                            spacing: 6px;
                        }
                        QRadioButton::indicator {
                            width: 12px;
                            height: 12px;
                            border-radius: 8px;
                            border: 2px solid #2196F3;
                            background-color: #1a1a1a;
                        }
                        QRadioButton::indicator:checked {
                            background-color: #2196F3;
                            border: 2px solid #2196F3;
                        }
                        QRadioButton::indicator:hover {
                            border: 2px solid #42A5F5;
                        }
                    """)

                    # Set random as default
                    random_radio.setChecked(True)

                    radio_layout.addWidget(random_radio)
                    radio_layout.addWidget(mutual_radio)
                    radio_layout.addStretch()

                    view_layout.addLayout(radio_layout)
                    options_layout.addLayout(view_layout)

                    # Store references
                    self.task_options[name] = {
                        'widget': options_widget,
                        'amount_input': amount_input,
                        'random_radio': random_radio,
                        'mutual_radio': mutual_radio
                    }

                # Connect arrow button to toggle options
                def make_toggle_func(arrow, opts, task_name):
                    def toggle():
                        is_visible = opts.isVisible()
                        opts.setVisible(not is_visible)
                        arrow.setText("▲" if not is_visible else "▼")
                    return toggle

                arrow_btn.clicked.connect(make_toggle_func(arrow_btn, options_widget, name))

            task_item_layout.addLayout(top_row)

            # Add options widget if exists
            if options_widget:
                task_item_layout.addWidget(options_widget)

            tasks_layout.addWidget(task_item)

            self.task_checkboxes.append({
                'checkbox': checkbox,
                'name': name,
                'icon': icon,
                'index': idx
            })

            if name == "View video":
                checkbox.setChecked(True)

        tasks_layout.addStretch()
        scroll.setWidget(tasks_container)
        layout.addWidget(scroll, alignment=Qt.AlignTop)

        # Connect radio button signals to toggle visibility
        def toggle_mode():
            is_tiktok = self.tiktok_radio.isChecked()
            scroll.setVisible(is_tiktok)
            # Set visibility for all 5 search time widgets
            for i in range(5):
                self.search_time_layout.itemAt(i).widget().setVisible(not is_tiktok)

        self.tiktok_radio.toggled.connect(toggle_mode)
        self.search_radio.toggled.connect(toggle_mode)

        toggle_mode()

        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(8)

        buttons_layout.addStretch()

        layout.addLayout(buttons_layout)

        # Add stretch to push everything to top
        layout.addStretch()

        group.setLayout(layout)
        return group

    def create_cookie_setting_section(self):
        group = QGroupBox("🍪 Cookie Setting")
        layout = QVBoxLayout()
        layout.setSpacing(8)
        layout.setContentsMargins(8, 8, 8, 8)

        # Cookie buttons grid (1 column)
        cookie_buttons_layout = QVBoxLayout()
        cookie_buttons_layout.setSpacing(6)

        view_cookie_btn = ModernButton("View files", "📂", "#9C27B0")
        view_cookie_btn.setMinimumHeight(32)
        view_cookie_btn.clicked.connect(self.view_cookie_folder)

        upload_cookie_btn = ModernButton("Upload files", "📥", "#FF5722")
        upload_cookie_btn.setMinimumHeight(32)
        upload_cookie_btn.clicked.connect(self.import_cookie_file)

        update_list_btn = ModernButton("Update to list", "🔄", "#4CAF50")
        update_list_btn.setMinimumHeight(32)
        update_list_btn.clicked.connect(self.update_cookie_list)

        import_profile_btn = ModernButton("Import to profile", "📥", "#FF9800")
        import_profile_btn.setMinimumHeight(32)
        import_profile_btn.clicked.connect(self.import_cookies_to_profiles)

        cookie_buttons_layout.addWidget(view_cookie_btn)
        cookie_buttons_layout.addWidget(upload_cookie_btn)
        cookie_buttons_layout.addWidget(update_list_btn)
        cookie_buttons_layout.addWidget(import_profile_btn)

        layout.addLayout(cookie_buttons_layout)

        group.setLayout(layout)
        return group

    def create_terminal_section(self):
        group = QGroupBox("🖥️ Terminal Output")
        group.setStyleSheet("""
            QGroupBox {
                background-color: #1a1a1a;
                border: 2px solid #2196F3;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 8px;
                font-weight: bold;
                font-size: 10pt;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                color: #2196F3;
            }
        """)
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(3)

        # Terminal container with absolute positioned buttons
        terminal_container = QFrame()
        terminal_container.setStyleSheet("background-color: transparent; border: none;")
        container_layout = QVBoxLayout(terminal_container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(0)

        # Terminal fills the container
        self.terminal = QTextEdit()
        self.terminal.setReadOnly(True)
        self.terminal.setFont(QFont("Consolas", 10))
        self.terminal.setMaximumHeight(100)
        container_layout.addWidget(self.terminal)

        sys.stdout = OutputRedirector(self.terminal, sys.__stdout__)
        sys.stderr = OutputRedirector(self.terminal, sys.__stderr__)

        # Absolute positioned buttons at bottom right (overlay on terminal)
        self.clear_btn = ModernButton("Clear", "🗑️", "#616161")
        self.clear_btn.setParent(terminal_container)
        self.clear_btn.setFixedSize(70, 28)
        self.clear_btn.setFont(QFont("Segoe UI", 8))
        self.clear_btn.clicked.connect(self.terminal.clear)
        self.clear_btn.raise_()  # Bring to front

        self.save_btn = ModernButton("Save", "💾", "#455A64")
        self.save_btn.setParent(terminal_container)
        self.save_btn.setFixedSize(70, 28)
        self.save_btn.setFont(QFont("Segoe UI", 8))
        self.save_btn.clicked.connect(self.save_log)
        self.save_btn.raise_()  # Bring to front

        # Override resize event to reposition buttons
        original_resize = terminal_container.resizeEvent
        def custom_resize(event):
            if original_resize:
                original_resize(event)
            self.position_terminal_buttons()
        terminal_container.resizeEvent = custom_resize

        layout.addWidget(terminal_container)
        group.setLayout(layout)

        # Position buttons initially
        QTimer.singleShot(100, self.position_terminal_buttons)

        return group

    def position_terminal_buttons(self):
        if hasattr(self, 'terminal') and hasattr(self, 'clear_btn') and hasattr(self, 'save_btn'):
            parent = self.clear_btn.parent()
            if parent:
                width = parent.width()
                height = parent.height()

                button_spacing = 6
                margin = 20

                save_x = width - self.save_btn.width() - margin
                save_y = height - self.save_btn.height() - margin/2

                clear_x = save_x - self.clear_btn.width() - button_spacing
                clear_y = save_y

                self.save_btn.move(save_x, save_y)
                self.clear_btn.move(clear_x, clear_y)

    @Slot()
    def browse_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File",
            config.EXCEL_PATH,
            "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.excel_path.setText(file_path)

    @Slot()
    def browse_cookies(self):
        dir_path = QFileDialog.getExistingDirectory(
            self, "Select Cookies Directory",
            config.COOKIES_DIR
        )
        if dir_path:
            self.cookies_dir.setText(dir_path)

    @Slot(int)
    def auto_save_threads(self, value):
        config.THREADS = value
        if hasattr(self, 'terminal'):
            self.log(f"⚙️ Threads updated to: {value}")

    @Slot(int)
    def auto_save_start_limit(self, value):
        config.START_LIMIT = value
        if hasattr(self, 'terminal'):
            self.log(f"🚀 Start Limit updated to: {value}")

    @Slot(int)
    def auto_save_full_screen(self, state):
        config.FULL_SCREEN = bool(state)
        if hasattr(self, 'terminal'):
            self.log(f"🖥️ Full screen mode: {'Enabled' if config.FULL_SCREEN else 'Disabled'}")

    @Slot()
    def view_cookie_folder(self):
        try:
            import os
            import subprocess
            import platform

            cookies_dir = config.COOKIES_DIR
            abs_path = os.path.abspath(cookies_dir)

            if not os.path.exists(abs_path):
                self.log(f"❌ Cookies folder not found: {abs_path}")
                return

            if platform.system() == 'Windows':
                subprocess.run(['explorer', abs_path], shell=False)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', abs_path])
            else:  # Linux
                subprocess.run(['xdg-open', abs_path])

            if hasattr(self, 'terminal'):
                self.log(f"📂 Opened cookies folder: {cookies_dir}")

        except Exception as e:
            error_msg = f"Failed to open cookies folder: {str(e)}"
            if hasattr(self, 'terminal'):
                self.log(f"❌ {error_msg}")

    @Slot()
    def import_cookie_file(self):
        try:
            import os
            import shutil

            cookies_dir = config.COOKIES_DIR

            os.makedirs(cookies_dir, exist_ok=True)

            # Open file dialog to select cookie files (multiple)
            file_paths, _ = QFileDialog.getOpenFileNames(
                self, "Select Cookie Files",
                "",
                "Text Files (*.txt);;All Files (*.*)"
            )

            if not file_paths:
                return

            imported_count = 0
            for file_path in file_paths:
                filename = os.path.basename(file_path)
                dest_path = os.path.join(cookies_dir, filename)

                # Skip if file already exists
                if os.path.exists(dest_path):
                    self.log(f"⚠️ Skipped existing file: {filename}")
                    continue

                shutil.copy2(file_path, dest_path)
                imported_count += 1

            if imported_count > 0:
                success_msg = f"✅ Imported {imported_count} cookie file(s)"
                if hasattr(self, 'terminal'):
                    self.log(success_msg)
            else:
                self.log("⚠️ No new files imported (all already exist)")

        except Exception as e:
            error_msg = f"Failed to import cookie files: {str(e)}"
            if hasattr(self, 'terminal'):
                self.log(f"❌ {error_msg}")

    @Slot()
    def update_cookie_list(self):
        try:
            import os
            from excel import update_excel_column_a_with_cookie_files

            cookies_dir = config.COOKIES_DIR
            excel_path = config.EXCEL_PATH

            if not os.path.exists(cookies_dir):
                error_msg = f"Cookie directory not found: {cookies_dir}"
                if hasattr(self, 'terminal'):
                    self.log(f"❌ {error_msg}")
                return

            if not os.path.exists(excel_path):
                error_msg = f"Excel file not found: {excel_path}"
                if hasattr(self, 'terminal'):
                    self.log(f"❌ {error_msg}")
                return

            count = update_excel_column_a_with_cookie_files(
                excel_path=excel_path,
                cookies_folder=cookies_dir
            )

            success_msg = f"✅ Updated Excel with {count} cookie files"
            if hasattr(self, 'terminal'):
                self.log(success_msg)

        except Exception as e:
            error_msg = f"Failed to update cookie list: {str(e)}"
            if hasattr(self, 'terminal'):
                self.log(f"❌ {error_msg}")

    def import_cookies_to_profiles(self):
        try:
            actions = {
                "create": False,
                "start": True,
                "import": True,
                "close": False,
                "delete": False,
                "pw": False,
                "handle_cookies": False
            }

            self.run_tasks_with_actions(actions, "Import Cookies to Profiles")

        except Exception as e:
            error_msg = f"Failed to import cookies to profiles: {str(e)}"
            if hasattr(self, 'terminal'):
                self.log(f"❌ {error_msg}")

    @Slot()
    def load_excel_data(self):
        try:
            import os
            excel_path = config.EXCEL_PATH

            if not os.path.exists(excel_path):
                self.profiles_text.setPlainText(f"❌ Excel file not found:\n{excel_path}")
                if hasattr(self, 'terminal'):
                    self.log(f"❌ Excel file not found: {excel_path}")
                return

            rows = read_excel()

            if not rows:
                self.profiles_text.setPlainText("")
                if hasattr(self, 'terminal'):
                    self.log("⚠️ No data found in Excel file")
                return

            # Extract profiles (column B only)
            profiles_list = []

            for row in rows:
                if len(row) >= 3:
                    # row is (profile_name, cookie_path, proxy)
                    proxy = row[2] if row[2] else ""    # proxy is at index 2
                    profiles_list.append(proxy)

            # Update text area
            profiles_text = "\n".join(profiles_list) if profiles_list else "⚠️ No profiles found"
            self.profiles_text.setPlainText(profiles_text)

            if hasattr(self, 'terminal'):
                self.log(f"✅ Loaded {len(rows)} rows from Excel: {excel_path}")

        except Exception as e:
            error_msg = f"❌ Error loading Excel data: {str(e)}"
            self.profiles_text.setPlainText(error_msg)
            if hasattr(self, 'terminal'):
                self.log(error_msg)

    def refresh_profiles_list(self):
        """Refresh the profiles list from GPM API"""
        try:
            self.profiles_list.clear()

            # Get profiles from GPM API
            profiles = get_profiles_list()

            if not profiles:
                item = QListWidgetItem("⚠️ No profiles found")
                item.setForeground(Qt.yellow)
                self.profiles_list.addItem(item)
                return

            # Add profiles to list
            for profile in profiles:
                # Create custom widget for profile item
                profile_item = ProfileListItem(profile['name'], profile['id'])
                profile_item.delete_clicked.connect(self.on_delete_profile)

                # Create list widget item and set the custom widget
                list_item = QListWidgetItem()
                list_item.setSizeHint(profile_item.sizeHint())
                self.profiles_list.addItem(list_item)
                self.profiles_list.setItemWidget(list_item, profile_item)

            if hasattr(self, 'terminal'):
                self.log(f"✅ Loaded {len(profiles)} profiles from GPM")

        except Exception as e:
            error_msg = f"❌ Error loading profiles: {str(e)}"
            self.profiles_list.clear()
            item = QListWidgetItem(error_msg)
            item.setForeground(Qt.red)
            self.profiles_list.addItem(item)
            if hasattr(self, 'terminal'):
                self.log(error_msg)

    def on_delete_profile(self, profile_id, profile_name):
        """Handle profile deletion"""
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Are you sure you want to delete the profile '{profile_name}'?\n\nThis action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                # Call delete API
                delete_profile(profile_id)

                # Log success
                if hasattr(self, 'terminal'):
                    self.log(f"✅ Successfully deleted profile: {profile_name}")

                # Refresh the profiles list
                self.refresh_profiles_list()

                # Show success message
                QMessageBox.information(
                    self,
                    "Success",
                    f"Profile '{profile_name}' has been deleted successfully."
                )

            except Exception as e:
                error_msg = f"❌ Failed to delete profile '{profile_name}': {str(e)}"
                if hasattr(self, 'terminal'):
                    self.log(error_msg)

                QMessageBox.critical(
                    self,
                    "Delete Failed",
                    f"Failed to delete profile '{profile_name}'.\n\nError: {str(e)}"
                )

    @Slot()
    def view_excel_file(self):
        try:
            import os
            import subprocess
            import platform

            excel_path = config.EXCEL_PATH

            abs_path = os.path.abspath(excel_path)

            if not os.path.exists(abs_path):
                self.log(f"❌ Excel file not found: {abs_path}")
                return

            # Open file with default application based on OS
            if platform.system() == 'Windows':
                # Use subprocess with 'start' command for Windows
                subprocess.run(['cmd', '/c', 'start', '', abs_path], shell=False)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', abs_path])
            else:  # Linux
                subprocess.run(['xdg-open', abs_path])

            if hasattr(self, 'terminal'):
                self.log(f"📊 Opened Excel file: {excel_path}")

        except Exception as e:
            error_msg = f"Failed to open Excel file: {str(e)}"
            if hasattr(self, 'terminal'):
                self.log(f"❌ {error_msg}")

    @Slot()
    def save_profiles_to_excel(self):
        try:
            from openpyxl import load_workbook
            import os

            excel_path = config.EXCEL_PATH

            if not os.path.exists(excel_path):
                self.log(f"❌ Excel file not found: {excel_path}")
                return

            # Get edited text from textarea
            text_content = self.profiles_text.toPlainText()
            lines = text_content.strip().split('\n')

            # Parse profiles from text - each line is a profile
            profiles_list = []
            for line in lines:
                line = line.strip()
                if line and not line.startswith('⚠️'):
                    profiles_list.append(line)

            # Allow saving empty list (clears all profiles)
            wb = load_workbook(excel_path)
            ws = wb.active

            # Clear all proxy cells from row 2 onwards
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=2).value = None

            # Write profiles in order from row 2
            updated_count = 0
            for i, proxy in enumerate(profiles_list, start=2):  # start=2 for row 2
                ws.cell(row=i, column=2).value = proxy
                updated_count += 1

            wb.save(excel_path)
            wb.close()

            if profiles_list:
                success_msg = f"✅ Saved {updated_count} profiles to Excel!"
            else:
                success_msg = "✅ Cleared all profiles from Excel!"
            if hasattr(self, 'terminal'):
                self.log(success_msg)

            # Reload to confirm changes
            self.load_excel_data()

        except Exception as e:
            error_msg = f"Failed to save profiles: {str(e)}"
            if hasattr(self, 'terminal'):
                self.log(f"❌ {error_msg}")

    @Slot()
    def apply_config(self):
        config.EXCEL_PATH = self.excel_path.text()
        config.COOKIES_DIR = self.cookies_dir.text()
        config.THREADS = self.threads_spin.value()
        config.START_LIMIT = self.start_limit.value()

        self.log("✅ Configuration updated successfully")

    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.terminal.append(f"[{timestamp}] {message}")

    def run_single_task(self, task_name, task_icon, task_index):
        if self.current_worker and self.current_worker.isRunning():
            self.log(f"⚠️ Another task is running. Please wait...")
            return

        # Reset stop flag
        self.stop_requested = False

        # Update stop button state (if it exists)
        if hasattr(self, 'stop_btn'):
            self.stop_btn.setEnabled(True)

        # Build actions dictionary for this single task
        actions = {
            "handle_cookies": False,
            "create": False,
            "start": False,
            "pw": False,
            "import": False,
            "close": False,
            "delete": False,
            "follow": False,
            "like_video": False,
            "comment": False,
            "view_video": False,
        }

        # Map task name to actions
        if task_name == "Start Profiles":
            actions["start"] = True
            # Check the "Run with CDP" checkbox to determine mode
            if hasattr(self, 'run_with_cdp_checkbox') and self.run_with_cdp_checkbox.isChecked():
                actions["pw"] = True
        elif task_name == "Close Profiles":
            actions["close"] = True
        elif task_name == "Delete Profiles":
            actions["delete"] = True
        elif task_name == "Follow":
            actions["follow"] = True
            actions["start"] = True
        elif task_name == "Like video":
            actions["like_video"] = True
            actions["start"] = True
        elif task_name == "Comment":
            actions["comment"] = True
            actions["start"] = True
        elif task_name == "View video":
            actions["view_video"] = True
            actions["start"] = True

        # Log the task being run
        self.log(f"\n{'='*70}")
        self.log(f"🚀 Running task: {task_icon} {task_name}")
        self.log(f"{'='*70}\n")

        # Run the task
        self.run_tasks_with_actions(actions)

    def run_single_task(self, task_name, task_icon, task_index):
        """Run a single automation task immediately"""
        if self.current_worker and self.current_worker.isRunning():
            self.log(f"⚠️ Another task is running. Please wait...")
            return

        # Reset stop flag
        self.stop_requested = False

        # Update stop button state if it exists
        if hasattr(self, 'stop_btn'):
            self.stop_btn.setEnabled(True)

        # Disable all task buttons except stop
        for task_item in self.task_buttons:
            task_item['button'].setEnabled(False)

        # Build actions dictionary for this single task
        actions = {
            "handle_cookies": False,
            "create": False,
            "start": False,
            "pw": False,
            "import": False,
            "close": False,
            "delete": False,
            "follow": False,
            "like_video": False,
            "comment": False,
            "view_video": False,
        }

        # Map task name to actions
        if task_name == "Start Profiles":
            actions["start"] = True
            # Check the "Run with CDP" checkbox to determine mode
            if hasattr(self, 'run_with_cdp_checkbox') and self.run_with_cdp_checkbox.isChecked():
                actions["pw"] = True
        elif task_name == "Close Profiles":
            actions["close"] = True
        elif task_name == "Delete Profiles":
            actions["delete"] = True

        # If starting profiles, also include selected behavior actions
        # LƯU Ý: Behavior actions CHỈ hoạt động với CDP mode
        if task_name == "Start Profiles":
            # Determine if we're in CDP mode based on checkbox
            is_cdp_mode = hasattr(self, 'run_with_cdp_checkbox') and self.run_with_cdp_checkbox.isChecked()

            # Check behavior mode (Tiktok or Search)
            behavior_mode = "tiktok" if self.tiktok_radio.isChecked() else "search"
            actions["behavior_mode"] = behavior_mode

            if behavior_mode == "search":
                # For search mode, get search time (hours * 60 + minutes)
                hours = self.search_hours_input.value()
                minutes = self.search_minutes_input.value()
                total_minutes = hours * 60 + minutes
                # Ensure minimum 1 minute
                actions["search_time"] = max(1, total_minutes)
            else:
                # For tiktok mode, collect behavior actions
                for checkbox_item in self.task_checkboxes:
                    if checkbox_item['checkbox'].isChecked():
                        name = checkbox_item['name']
                        if name == "Follow":
                            actions["follow"] = True
                            # Lấy follow mode và target username từ GUI
                            if "Follow" in self.task_options:
                                follow_opts = self.task_options["Follow"]
                                if follow_opts['random_radio'].isChecked():
                                    actions["follow_mode"] = "random"
                                elif follow_opts['mutual_radio'].isChecked():
                                    actions["follow_mode"] = "mutual"
                                elif follow_opts['target_radio'].isChecked():
                                    actions["follow_mode"] = "target"
                                    target_text = follow_opts['target_input'].text().strip()
                                    if target_text:
                                        actions["follow_target"] = target_text
                        elif name == "View video":
                            actions["view_video"] = True
                            # Lấy view amount từ GUI nếu có
                            if "View video" in self.task_options:
                                view_opts = self.task_options["View video"]
                                actions["view_amount"] = view_opts['amount_input'].value()
                        elif name == "Like video":
                            actions["like_video"] = True
                        elif name == "Comment":
                            actions["comment"] = True

            # CHỈ enable pw mode nếu:
            # 1. Đang ở CDP mode (checkbox checked)
            # 2. VÀ có ít nhất một behavior action được chọn (tiktok mode) HOẶC search mode được chọn
            if is_cdp_mode:
                has_behavior_actions = actions["follow"] or actions["like_video"] or actions["comment"] or actions["view_video"]
                if has_behavior_actions or behavior_mode == "search":
                    actions["pw"] = True
            else:
                # Nếu ở launch mode mà có behavior actions được chọn, cảnh báo user
                has_behavior_actions = actions["follow"] or actions["like_video"] or actions["comment"] or actions["view_video"]
                if has_behavior_actions or behavior_mode == "search":
                    self.log("⚠️ WARNING: Behavior actions và Search mode chỉ hoạt động với CDP mode!")
                    self.log("⚠️ Vui lòng bật checkbox 'Run with CDP' để sử dụng các tính năng này.")
                    self.log("⚠️ Hiện tại chỉ start profiles mà không chạy behavior actions hoặc search.\n")

        # Log the task being run
        self.log(f"\n{'='*70}")
        self.log(f"🚀 Running task: {task_icon} {task_name}")
        self.log(f"{'='*70}\n")

        # Run the task
        self.run_tasks_with_actions(actions, f"{task_icon} {task_name}")


    def stop_tasks(self):
        """Stop currently running tasks"""
        if not self.current_worker or not self.current_worker.isRunning():
            self.log("⚠️ No tasks are currently running.")
            return

        self.log("\n🛑 Stopping tasks... Please wait...")
        self.stop_requested = True

        # Request worker to stop
        if self.current_worker:
            self.current_worker.stop()
            # Wait for worker to finish (with timeout)
            if self.current_worker.wait(5000):  # 5 seconds timeout
                self.log("✅ Tasks stopped successfully")
            else:
                self.log("⚠️ Force terminating tasks...")
                self.current_worker.terminate()
                self.current_worker.wait()

        # Reset button states
        self.stop_btn.setEnabled(False)

        # Re-enable all task buttons
        for task_item in self.task_buttons:
            task_item['button'].setEnabled(True)

    def run_tasks_with_actions(self, actions, task_name="Tasks"):
        """Run tasks using CLI logic with ThreadPoolExecutor"""
        def run():
            from concurrent.futures import ThreadPoolExecutor, as_completed
            import threading

            # Check if stop was requested
            if self.stop_requested or (self.current_worker and self.current_worker.is_stopped()):
                safe_print("⏹️ Task execution stopped by user")
                return

            # Initialize semaphore for start limit
            config.start_sem = threading.Semaphore(config.START_LIMIT)

            # Clear previous state
            with config.started_lock:
                config.started_debug_addrs.clear()

            with config.pw_jobs_lock:
                config.pw_jobs.clear()

            # Read excel data
            rows = read_excel()

            # Calculate total profiles for dynamic grid
            total_profiles = len(rows) if (actions["start"] or actions["import"] or actions["pw"]) else None

            # Execute with thread pool (same as CLI)
            with ThreadPoolExecutor(max_workers=config.THREADS) as ex:
                futures = [
                    ex.submit(process_row, *row[:3], i+1, actions, total_profiles)
                    for i, row in enumerate(rows)
                ]
                for future in as_completed(futures):
                    # Check stop request periodically
                    if self.stop_requested or (self.current_worker and self.current_worker.is_stopped()):
                        safe_print("⏹️ Stopping task execution...")
                        # Cancel remaining futures
                        for f in futures:
                            f.cancel()
                        ex.shutdown(wait=False)
                        return

            # Check before running playwright jobs
            if self.stop_requested or (self.current_worker and self.current_worker.is_stopped()):
                safe_print("⏹️ Task execution stopped before Playwright jobs")
                return

            # Run playwright jobs if needed
            if actions["import"] or actions["pw"]:
                with config.pw_jobs_lock:
                    jobs = config.pw_jobs.copy()

                if not jobs:
                    safe_print("⚠️ No playwright jobs collected.")
                else:
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    try:
                        loop.run_until_complete(run_all_playwright(jobs, actions))
                    except Exception as e:
                        if not self.stop_requested:
                            safe_print(f"❌ Playwright error: {e}")
                    finally:
                        loop.close()

            if not self.stop_requested and not (self.current_worker and self.current_worker.is_stopped()):
                safe_print("✅ ALL DONE")
            else:
                safe_print("⏹️ Tasks stopped by user")

        self.current_worker = TaskWorker(task_name, run)
        self.current_worker.log_signal.connect(self.log)
        self.current_worker.finished_signal.connect(self.on_tasks_finished)
        self.current_worker.start()

    @Slot(bool, str)
    def on_tasks_finished(self, success, message):
        if hasattr(self, 'stop_btn'):
            self.stop_btn.setEnabled(False)

        # Re-enable all task buttons
        for task_item in self.task_buttons:
            task_item['button'].setEnabled(True)

        if success:
            self.log("\n✅ Task completed successfully!")
        else:
            if "stopped" in message.lower():
                self.log(f"\n⏹️ {message}")
            else:
                self.log(f"\n❌ Task failed: {message}")

    def save_log(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Log",
            f"gpm_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            "Text Files (*.txt)"
        )
        if file_path:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(self.terminal.toPlainText())
            self.log(f"💾 Log saved: {file_path}")

    def create_profiles(self):
        if self.current_worker and self.current_worker.isRunning():
            self.log("⚠️ Another task is running. Please wait...")
            return

        self.stop_requested = False

        if hasattr(self, 'stop_btn'):
            self.stop_btn.setEnabled(True)

        actions = {
            "handle_cookies": False,
            "create": True,
            "start": False,
            "pw": False,
            "import": False,
            "close": False,
            "delete": False,
        }

        self.log(f"\n{'='*70}")
        self.log("📋 Running task: ➕ Create Profiles")
        self.log(f"{'='*70}\n")

        self.run_tasks_with_actions(actions, "➕ Create Profiles")

    def update_profiles(self):
        if self.current_worker and self.current_worker.isRunning():
            self.log("⚠️ Another task is running. Please wait...")
            return

        self.stop_requested = False

        if hasattr(self, 'stop_btn'):
            self.stop_btn.setEnabled(True)

        actions = {
            "handle_cookies": False,
            "create": False,
            "start": False,
            "pw": False,
            "import": False,
            "close": False,
            "delete": False,
            "update": True,
        }

        self.log(f"\n{'='*70}")
        self.log("📋 Running task: 🔄 Update Profiles")
        self.log(f"{'='*70}\n")

        self.run_tasks_with_actions(actions, "🔄 Update Profiles")

def run_gui():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    window = GPMMainWindow()
    window.show()

    sys.exit(app.exec())

def main():
    try:
        run_gui()
    except Exception as e:
        print(f"❌ Failed to launch GUI: {e}")

if __name__ == "__main__":
    main()