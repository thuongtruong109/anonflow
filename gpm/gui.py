"""
GPM Automation - GUI Application
Modern Qt-based interface for GPM automation tasks
"""
import sys
import asyncio
import threading
from pathlib import Path
from datetime import datetime
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTextEdit, QLabel, QGroupBox, QGridLayout, QMessageBox,
    QFileDialog, QSplitter, QFrame, QLineEdit, QSpinBox, QCheckBox,
    QScrollArea
)
from PySide6.QtCore import Qt, QThread, Signal, Slot, QPropertyAnimation, QEasingCurve, QTimer
from PySide6.QtGui import QFont, QTextCursor

# Import GPM modules
import config
from config import setup_logger
from excel import update_excel_column_a_with_cookie_files, read_excel
from services import create_profile, start_profile, close_profile, delete_profile
from runner import run_all_playwright
from utils import safe_print, menu_multi_select
from index import process_row


class OutputRedirector:
    """Redirect stdout/stderr to GUI terminal"""
    def __init__(self, text_widget, original_stream):
        self.text_widget = text_widget
        self.original_stream = original_stream

    def write(self, text):
        if text.strip():
            # Also write to original stream for debugging
            if self.original_stream:
                self.original_stream.write(text)

            # Write to GUI
            cursor = self.text_widget.textCursor()
            cursor.movePosition(QTextCursor.End)
            self.text_widget.setTextCursor(cursor)
            self.text_widget.insertPlainText(text)
            self.text_widget.ensureCursorVisible()

    def flush(self):
        if self.original_stream:
            self.original_stream.flush()


class TaskWorker(QThread):
    """Worker thread for running async tasks"""
    log_signal = Signal(str)
    finished_signal = Signal(bool, str)

    def __init__(self, task_name, task_func):
        super().__init__()
        self.task_name = task_name
        self.task_func = task_func
        self._is_stopped = False

    def stop(self):
        """Request the worker to stop"""
        self._is_stopped = True

    def is_stopped(self):
        """Check if stop was requested"""
        return self._is_stopped

    def run(self):
        try:
            self.log_signal.emit(f"\n{'='*70}\n🚀 Starting: {self.task_name}\n{'='*70}\n")

            # Run the task
            self.task_func()

            if self._is_stopped:
                self.log_signal.emit(f"\n{'='*70}\n⏹️ Stopped: {self.task_name}\n{'='*70}\n")
                self.finished_signal.emit(False, f"Task '{self.task_name}' was stopped by user")
            else:
                self.log_signal.emit(f"\n{'='*70}\n✅ Completed: {self.task_name}\n{'='*70}\n")
                self.finished_signal.emit(True, f"Task '{self.task_name}' completed!")

        except Exception as e:
            error_msg = f"Error: {str(e)}"
            self.log_signal.emit(f"\n{'='*70}\n❌ Failed: {self.task_name}\n{error_msg}\n{'='*70}\n")
            self.finished_signal.emit(False, error_msg)


class ModernButton(QPushButton):
    """Styled button with modern look"""
    def __init__(self, text, icon="", color="#2196F3"):
        super().__init__()
        self.base_color = color
        display_text = f"{icon}  {text}" if icon else text
        self.setText(display_text)
        self.setMinimumHeight(35)
        self.setFont(QFont("Segoe UI", 9, QFont.Bold))
        self.setCursor(Qt.PointingHandCursor)
        self.update_style()

    def update_style(self):
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {self.base_color};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 12px;
                text-align: left;
            }}
            QPushButton:hover {{
                background-color: {self.lighten_color(self.base_color, 20)};
            }}
            QPushButton:pressed {{
                background-color: {self.lighten_color(self.base_color, -20)};
            }}
            QPushButton:disabled {{
                background-color: #424242;
                color: #757575;
            }}
        """)

    def lighten_color(self, hex_color, amount):
        hex_color = hex_color.lstrip('#')
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        r = max(0, min(255, r + amount))
        g = max(0, min(255, g + amount))
        b = max(0, min(255, b + amount))
        return f"#{r:02x}{g:02x}{b:02x}"

class GPMMainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.current_worker = None
        self.logger = setup_logger("gui")
        self.task_checkboxes = []
        self.task_queue = []
        self.task_options = {}
        self.stop_requested = False
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Anonflow")
        self.setMinimumSize(1280, 750)

        self.apply_dark_theme()

        # Main widget
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        # Top row: Header + Settings (left) and Tasks (right) with 3:7 ratio
        top_row_layout = QHBoxLayout()
        top_row_layout.setSpacing(10)

        # Left column (30%): Header + Threads/Start Limit
        left_column = QVBoxLayout()
        left_column.setSpacing(10)
        left_column.addWidget(self.create_header())
        left_column.addWidget(self.create_quick_settings())

        # Create a widget to hold the left column
        left_widget = QWidget()
        left_widget.setLayout(left_column)

        # Add to top row with ratio 2:8
        top_row_layout.addWidget(left_widget, 2)
        top_row_layout.addWidget(self.create_config_section(), 8)  # Excel Data Preview
        layout.addLayout(top_row_layout)

        # Main splitter
        splitter = QSplitter(Qt.Vertical)

        # Top section: Tasks
        top_widget = QWidget()
        top_layout = QHBoxLayout(top_widget)
        top_layout.setSpacing(10)
        top_layout.addWidget(self.create_tasks_section(), 1)
        splitter.addWidget(top_widget)

        # Bottom section: Terminal
        splitter.addWidget(self.create_terminal_section())

        # Increase task section height, decrease terminal height
        splitter.setSizes([550, 200])
        layout.addWidget(splitter)

    def apply_dark_theme(self):
        self.setStyleSheet("""
            QMainWindow, QWidget {
                background-color: #0d0d0d;
                color: #e0e0e0;
                font-family: 'Segoe UI', Arial;
            }
            QGroupBox {
                background-color: #1a1a1a;
                border: 2px solid #2196F3;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                font-weight: bold;
                font-size: 10pt;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                color: #2196F3;
            }
            QTextEdit {
                background-color: #000000;
                color: #00ff00;
                border: 2px solid #2196F3;
                border-radius: 6px;
                padding: 8px;
                font-family: 'Consolas', 'Courier New', monospace;
                font-size: 9pt;
                selection-background-color: #2196F3;
            }
            QLineEdit {
                background-color: #1a1a1a;
                color: #ffffff;
                border: 1px solid #333;
                border-radius: 4px;
                padding: 6px;
                font-size: 9pt;
            }
            QLineEdit:focus {
                border: 1px solid #2196F3;
            }
            QSpinBox {
                background-color: #1a1a1a;
                color: #ffffff;
                border: 1px solid #333;
                border-radius: 4px;
                padding: 4px;
                font-size: 9pt;
            }
            QSpinBox:focus {
                border: 1px solid #2196F3;
            }
            QCheckBox {
                spacing: 5px;
                color: #e0e0e0;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
                border-radius: 3px;
                border: 1px solid #2196F3;
            }
            QCheckBox::indicator:checked {
                background-color: #2196F3;
            }
            QLabel {
                color: #e0e0e0;
            }
            QSplitter::handle {
                background-color: transparent;
                height: 0px;
            }
        """)

    def create_header(self):
        frame = QFrame()
        frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #1a237e, stop:1 #2196F3);
                border-radius: 8px;
                padding: 8px 12px;
            }
        """)
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(1)

        title = QLabel("🚀 GPM Automation Suite")
        title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title.setStyleSheet("color: white;")

        subtitle = QLabel("TikTok Profile & Cookie Management")
        subtitle.setFont(QFont("Segoe UI", 8))
        subtitle.setStyleSheet("color: #bbdefb;")

        layout.addWidget(title)
        layout.addWidget(subtitle)
        return frame

    def create_quick_settings(self):
        """Create quick settings section for Threads and Start Limit"""
        group = QGroupBox("⚙️ Quick Settings")
        layout = QVBoxLayout()
        layout.setSpacing(8)
        layout.setContentsMargins(8, 8, 8, 8)

        # Threads and Start Limit on same row
        settings_layout = QHBoxLayout()
        settings_layout.setSpacing(10)

        # Threads
        threads_label = QLabel("Threads:")
        threads_label.setFont(QFont("Segoe UI", 9))
        self.threads_spin = QSpinBox()
        self.threads_spin.setRange(1, 20)
        self.threads_spin.setValue(config.THREADS)
        self.threads_spin.valueChanged.connect(self.auto_save_threads)
        self.threads_spin.setMaximumWidth(80)

        settings_layout.addWidget(threads_label)
        settings_layout.addWidget(self.threads_spin)

        # Start Limit
        limit_label = QLabel("Start Limit:")
        limit_label.setFont(QFont("Segoe UI", 9))
        self.start_limit = QSpinBox()
        self.start_limit.setRange(1, 50)
        self.start_limit.setValue(config.START_LIMIT)
        self.start_limit.valueChanged.connect(self.auto_save_start_limit)
        self.start_limit.setMaximumWidth(80)

        settings_layout.addWidget(limit_label)
        settings_layout.addWidget(self.start_limit)
        settings_layout.addStretch()

        layout.addLayout(settings_layout)

        # Cookie buttons row
        cookie_buttons_layout = QHBoxLayout()
        cookie_buttons_layout.setSpacing(6)

        view_cookie_btn = ModernButton("View Cookies", "📂", "#9C27B0")
        view_cookie_btn.setMinimumHeight(32)
        view_cookie_btn.clicked.connect(self.view_cookie_folder)

        import_cookie_btn = ModernButton("Import Cookie", "📥", "#FF5722")
        import_cookie_btn.setMinimumHeight(32)
        import_cookie_btn.clicked.connect(self.import_cookie_file)

        cookie_buttons_layout.addWidget(view_cookie_btn)
        cookie_buttons_layout.addWidget(import_cookie_btn)

        layout.addLayout(cookie_buttons_layout)

        group.setLayout(layout)
        return group

    def create_config_section(self):
        group = QGroupBox("⚙️ Proxy Setting")
        group.setStyleSheet("""
            QGroupBox {
                background-color: #1a1a1a;
                border: 2px solid #2196F3;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                font-weight: bold;
                font-size: 10pt;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                color: #2196F3;
            }
        """)
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(8, 8, 8, 8)

        self.proxies_text = QTextEdit()
        self.proxies_text.setReadOnly(False)
        self.proxies_text.setFont(QFont("Consolas", 8))
        self.proxies_text.setStyleSheet("""
            QTextEdit {
                background-color: #0a0a0a;
                color: #ffaa00;
                border: 1px solid #2196F3;
                border-radius: 4px;
                padding: 6px;
                font-family: 'Consolas', 'Courier New', monospace;
            }
            QTextEdit:focus {
                border: 2px solid #42A5F5;
            }
        """)
        self.proxies_text.setPlaceholderText("Proxy addresses will be displayed here...\nEdit and save to update Excel file.")

        main_layout.addWidget(self.proxies_text)

        # Bottom: Action buttons
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(8)

        reload_btn = ModernButton("Reload Excel Data", "🔄", "#FF9800")
        reload_btn.clicked.connect(self.load_excel_data)

        view_btn = ModernButton("View Data File", "📊", "#9C27B0")
        view_btn.clicked.connect(self.view_excel_file)

        save_btn = ModernButton("Save Changes", "💾", "#4CAF50")
        save_btn.clicked.connect(self.save_proxies_to_excel)

        buttons_layout.addWidget(reload_btn)
        buttons_layout.addWidget(view_btn)
        buttons_layout.addWidget(save_btn)

        main_layout.addLayout(buttons_layout)

        group.setLayout(main_layout)

        self.load_excel_data()

        return group

    def create_tasks_section(self):
        group = QGroupBox("📋 Automation Tasks")
        layout = QVBoxLayout()
        layout.setSpacing(8)
        layout.setContentsMargins(8, 8, 8, 8)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                border: none;
                background-color: #1a1a1a;
                width: 10px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical {
                background-color: #2196F3;
                border-radius: 5px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #42A5F5;
            }
        """)

        # Container for all tasks
        tasks_container = QWidget()
        tasks_layout = QVBoxLayout(tasks_container)
        tasks_layout.setSpacing(6)
        tasks_layout.setContentsMargins(0, 0, 0, 0)

        # Define tasks with their metadata (name, icon, original_index, has_options)
        tasks = [
            ("Update Cookies to Excel", "📝", 0, False),
            ("Create Profiles", "➕", 1, False),
            ("Import Cookies", "📥", 2, False),
            ("Start Profiles", "▶️", 3, False),
            ("Start Profiles (with CDP)", "🎬", 4, False),
            ("Close Profiles", "⏹️", 5, False),
            ("Delete Profiles", "🗑️", 6, False),
        ]

        self.task_checkboxes = []
        self.task_options = {}  # Store option widgets for each task

        for i, (name, icon, idx, has_options) in enumerate(tasks):
            # Create task item container
            task_item = QFrame()
            task_item.setStyleSheet("""
                QFrame {
                    background-color: #1a1a1a;
                    border: 1px solid #333;
                    border-radius: 6px;
                    padding: 4px;
                }
                QFrame:hover {
                    border: 1px solid #2196F3;
                }
            """)
            task_item_layout = QVBoxLayout(task_item)
            task_item_layout.setContentsMargins(4, 4, 4, 4)
            task_item_layout.setSpacing(4)

            # Top row: Checkbox and arrow button
            top_row = QHBoxLayout()
            top_row.setSpacing(4)

            checkbox = QCheckBox(f"{icon}  {name}")
            checkbox.setFont(QFont("Segoe UI", 9))
            checkbox.setStyleSheet("""
                QCheckBox {
                    spacing: 8px;
                    color: #e0e0e0;
                    padding: 4px;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                    border-radius: 3px;
                    border: 2px solid #2196F3;
                    background-color: #1a1a1a;
                }
                QCheckBox::indicator:checked {
                    background-color: #2196F3;
                    border: 2px solid #2196F3;
                }
                QCheckBox::indicator:hover {
                    border: 2px solid #42A5F5;
                }
                QCheckBox:hover {
                    color: #42A5F5;
                }
            """)

            top_row.addWidget(checkbox)

            # Add arrow button if task has options
            arrow_btn = None
            options_widget = None
            if has_options:
                arrow_btn = QPushButton("▼")
                arrow_btn.setMaximumWidth(30)
                arrow_btn.setMaximumHeight(30)
                arrow_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        font-size: 12px;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background-color: #42A5F5;
                    }
                    QPushButton:pressed {
                        background-color: #1976D2;
                    }
                """)
                arrow_btn.setCursor(Qt.PointingHandCursor)
                top_row.addWidget(arrow_btn)

                # Create options widget (initially hidden)
                options_widget = QWidget()
                options_widget.setVisible(False)
                options_layout = QVBoxLayout(options_widget)
                options_layout.setContentsMargins(20, 4, 4, 4)
                options_layout.setSpacing(4)

                # Add specific options based on task
                if name == "Run Automation":
                    # Add follower input as example
                    follower_layout = QHBoxLayout()
                    follower_label = QLabel("👥 Target Followers:")
                    follower_label.setFont(QFont("Segoe UI", 9))
                    follower_label.setStyleSheet("color: #bbdefb;")

                    follower_input = QSpinBox()
                    follower_input.setRange(1, 10000)
                    follower_input.setValue(100)
                    follower_input.setStyleSheet("""
                        QSpinBox {
                            background-color: #0d0d0d;
                            color: #2196F3;
                            border: 1px solid #2196F3;
                            border-radius: 4px;
                            padding: 4px;
                            font-size: 9pt;
                            font-weight: bold;
                        }
                        QSpinBox:focus {
                            border: 2px solid #42A5F5;
                        }
                    """)
                    follower_input.setMinimumWidth(100)

                    follower_layout.addWidget(follower_label)
                    follower_layout.addWidget(follower_input)
                    follower_layout.addStretch()

                    options_layout.addLayout(follower_layout)

                    # Store reference to follower input
                    self.task_options[name] = {
                        'widget': options_widget,
                        'follower_input': follower_input
                    }

                # Connect arrow button to toggle options
                def make_toggle_func(arrow, opts, task_name):
                    def toggle():
                        is_visible = opts.isVisible()
                        opts.setVisible(not is_visible)
                        arrow.setText("▲" if not is_visible else "▼")
                    return toggle

                arrow_btn.clicked.connect(make_toggle_func(arrow_btn, options_widget, name))

            task_item_layout.addLayout(top_row)

            # Add options widget if exists
            if options_widget:
                task_item_layout.addWidget(options_widget)

            tasks_layout.addWidget(task_item)

            self.task_checkboxes.append({
                'checkbox': checkbox,
                'name': name,
                'icon': icon,
                'index': idx
            })

        tasks_layout.addStretch()
        scroll.setWidget(tasks_container)
        layout.addWidget(scroll)

        # Buttons layout for Run and Stop
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(8)

        self.run_btn = ModernButton("Run Selected Tasks", "🚀", "#4CAF50")
        self.run_btn.setMinimumHeight(35)
        self.run_btn.setFont(QFont("Segoe UI", 10, QFont.Bold))
        self.run_btn.clicked.connect(self.run_selected_tasks)

        self.stop_btn = ModernButton("Stop Tasks", "🛑", "#F44336")
        self.stop_btn.setMinimumHeight(35)
        self.stop_btn.setFont(QFont("Segoe UI", 10, QFont.Bold))
        self.stop_btn.clicked.connect(self.stop_tasks)
        self.stop_btn.setEnabled(False)  # Disabled by default

        buttons_layout.addWidget(self.run_btn, 2)
        buttons_layout.addWidget(self.stop_btn, 1)

        layout.addLayout(buttons_layout)

        group.setLayout(layout)
        return group

    def create_terminal_section(self):
        group = QGroupBox("🖥️ Terminal Output")
        group.setStyleSheet("""
            QGroupBox {
                background-color: #1a1a1a;
                border: 2px solid #2196F3;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                font-weight: bold;
                font-size: 10pt;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                color: #2196F3;
            }
        """)
        layout = QVBoxLayout()
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(3)

        # Terminal container with absolute positioned buttons
        terminal_container = QFrame()
        terminal_container.setStyleSheet("background-color: transparent; border: none;")
        container_layout = QVBoxLayout(terminal_container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(0)

        # Terminal fills the container
        self.terminal = QTextEdit()
        self.terminal.setReadOnly(True)
        self.terminal.setFont(QFont("Consolas", 10))
        container_layout.addWidget(self.terminal)

        sys.stdout = OutputRedirector(self.terminal, sys.__stdout__)
        sys.stderr = OutputRedirector(self.terminal, sys.__stderr__)

        # Absolute positioned buttons at bottom right (overlay on terminal)
        self.clear_btn = ModernButton("Clear", "🗑️", "#616161")
        self.clear_btn.setParent(terminal_container)
        self.clear_btn.setFixedSize(70, 28)
        self.clear_btn.setFont(QFont("Segoe UI", 8))
        self.clear_btn.clicked.connect(self.terminal.clear)
        self.clear_btn.raise_()  # Bring to front

        self.save_btn = ModernButton("Save", "💾", "#455A64")
        self.save_btn.setParent(terminal_container)
        self.save_btn.setFixedSize(70, 28)
        self.save_btn.setFont(QFont("Segoe UI", 8))
        self.save_btn.clicked.connect(self.save_log)
        self.save_btn.raise_()  # Bring to front

        # Override resize event to reposition buttons
        original_resize = terminal_container.resizeEvent
        def custom_resize(event):
            if original_resize:
                original_resize(event)
            self.position_terminal_buttons()
        terminal_container.resizeEvent = custom_resize

        layout.addWidget(terminal_container)
        group.setLayout(layout)

        # Position buttons initially
        QTimer.singleShot(100, self.position_terminal_buttons)

        return group

    def position_terminal_buttons(self):
        if hasattr(self, 'terminal') and hasattr(self, 'clear_btn') and hasattr(self, 'save_btn'):
            parent = self.clear_btn.parent()
            if parent:
                width = parent.width()
                height = parent.height()

                button_spacing = 6
                margin = 10

                save_x = width - self.save_btn.width() - margin
                save_y = height - self.save_btn.height() - margin

                clear_x = save_x - self.clear_btn.width() - button_spacing
                clear_y = save_y

                self.save_btn.move(save_x, save_y)
                self.clear_btn.move(clear_x, clear_y)

    @Slot()
    def browse_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File",
            config.EXCEL_PATH,
            "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.excel_path.setText(file_path)

    @Slot()
    def browse_cookies(self):
        dir_path = QFileDialog.getExistingDirectory(
            self, "Select Cookies Directory",
            config.COOKIES_DIR
        )
        if dir_path:
            self.cookies_dir.setText(dir_path)

    @Slot(int)
    def auto_save_threads(self, value):
        config.THREADS = value
        if hasattr(self, 'terminal'):
            self.log(f"⚙️ Threads updated to: {value}")

    @Slot(int)
    def auto_save_start_limit(self, value):
        config.START_LIMIT = value
        if hasattr(self, 'terminal'):
            self.log(f"🚀 Start Limit updated to: {value}")

    @Slot()
    def view_cookie_folder(self):
        try:
            import os
            import subprocess
            import platform

            cookies_dir = config.COOKIES_DIR
            abs_path = os.path.abspath(cookies_dir)

            if not os.path.exists(abs_path):
                self.log(f"❌ Cookies folder not found: {abs_path}")
                return

            if platform.system() == 'Windows':
                subprocess.run(['explorer', abs_path], shell=False)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', abs_path])
            else:  # Linux
                subprocess.run(['xdg-open', abs_path])

            if hasattr(self, 'terminal'):
                self.log(f"📂 Opened cookies folder: {cookies_dir}")

        except Exception as e:
            error_msg = f"Failed to open cookies folder: {str(e)}"
            if hasattr(self, 'terminal'):
                self.log(f"❌ {error_msg}")

    @Slot()
    def import_cookie_file(self):
        try:
            import os
            import shutil

            cookies_dir = config.COOKIES_DIR

            os.makedirs(cookies_dir, exist_ok=True)

            # Open file dialog to select cookie file
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Select Cookie File",
                "",
                "JSON Files (*.json);;All Files (*.*)"
            )

            if not file_path:
                return

            filename = os.path.basename(file_path)
            dest_path = os.path.join(cookies_dir, filename)

            shutil.copy2(file_path, dest_path)

            success_msg = f"✅ Cookie file imported: {filename}"
            if hasattr(self, 'terminal'):
                self.log(success_msg)

        except Exception as e:
            error_msg = f"Failed to import cookie file: {str(e)}"
            if hasattr(self, 'terminal'):
                self.log(f"❌ {error_msg}")    @Slot()
    def load_excel_data(self):
        try:
            import os
            excel_path = config.EXCEL_PATH

            if not os.path.exists(excel_path):
                self.proxies_text.setPlainText(f"❌ Excel file not found:\n{excel_path}")
                if hasattr(self, 'terminal'):
                    self.log(f"❌ Excel file not found: {excel_path}")
                return

            rows = read_excel()

            if not rows:
                self.proxies_text.setPlainText("⚠️ No data found in Excel file")
                if hasattr(self, 'terminal'):
                    self.log("⚠️ No data found in Excel file")
                return

            # Extract proxies (column B only)
            proxies_list = []

            for i, row in enumerate(rows, 1):
                if len(row) >= 3:
                    # row is (profile_name, cookie_path, proxy)
                    proxy = row[2] if row[2] else ""    # proxy is at index 2
                    proxies_list.append(f"{i}. {proxy}")

            # Update text area
            proxies_text = "\n".join(proxies_list) if proxies_list else "⚠️ No proxies found"
            self.proxies_text.setPlainText(proxies_text)

            if hasattr(self, 'terminal'):
                self.log(f"✅ Loaded {len(rows)} rows from Excel: {excel_path}")

        except Exception as e:
            error_msg = f"❌ Error loading Excel data: {str(e)}"
            self.proxies_text.setPlainText(error_msg)
            if hasattr(self, 'terminal'):
                self.log(error_msg)

    @Slot()
    def view_excel_file(self):
        try:
            import os
            import subprocess
            import platform

            excel_path = config.EXCEL_PATH

            abs_path = os.path.abspath(excel_path)

            if not os.path.exists(abs_path):
                self.log(f"❌ Excel file not found: {abs_path}")
                return

            # Open file with default application based on OS
            if platform.system() == 'Windows':
                # Use subprocess with 'start' command for Windows
                subprocess.run(['cmd', '/c', 'start', '', abs_path], shell=False)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', abs_path])
            else:  # Linux
                subprocess.run(['xdg-open', abs_path])

            if hasattr(self, 'terminal'):
                self.log(f"📊 Opened Excel file: {excel_path}")

        except Exception as e:
            error_msg = f"Failed to open Excel file: {str(e)}"
            if hasattr(self, 'terminal'):
                self.log(f"❌ {error_msg}")

    @Slot()
    def save_proxies_to_excel(self):
        try:
            from openpyxl import load_workbook
            import os

            excel_path = config.EXCEL_PATH

            if not os.path.exists(excel_path):
                self.log(f"❌ Excel file not found: {excel_path}")
                return

            # Get edited text from textarea
            text_content = self.proxies_text.toPlainText()
            lines = text_content.strip().split('\n')

            # Parse proxies from text (format: "1. proxy_address")
            proxies_dict = {}
            for line in lines:
                line = line.strip()
                if line and '. ' in line:
                    try:
                        index_str, proxy = line.split('. ', 1)
                        index = int(index_str)
                        proxies_dict[index] = proxy.strip()
                    except ValueError:
                        continue

            if not proxies_dict:
                self.log("⚠️ No valid proxy data to save. Format: 1. proxy_address")
                return

            wb = load_workbook(excel_path)
            ws = wb.active

            updated_count = 0
            for row_num, proxy in proxies_dict.items():
                excel_row = row_num + 1  # +1 because Excel rows start at 1, +1 for header
                if excel_row <= ws.max_row:
                    ws.cell(excel_row, 2).value = proxy  # Column B (proxy)
                    updated_count += 1

            wb.save(excel_path)
            wb.close()

            success_msg = f"✅ Saved {updated_count} proxies to Excel!"
            if hasattr(self, 'terminal'):
                self.log(success_msg)

            # Reload to confirm changes
            self.load_excel_data()

        except Exception as e:
            error_msg = f"Failed to save proxies: {str(e)}"
            if hasattr(self, 'terminal'):
                self.log(f"❌ {error_msg}")

    @Slot()
    def apply_config(self):
        config.EXCEL_PATH = self.excel_path.text()
        config.COOKIES_DIR = self.cookies_dir.text()
        config.THREADS = self.threads_spin.value()
        config.START_LIMIT = self.start_limit.value()

        self.log("✅ Configuration updated successfully")

    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.terminal.append(f"[{timestamp}] {message}")

    def run_selected_tasks(self):
        """Run all selected tasks - synchronized with CLI logic"""
        if self.current_worker and self.current_worker.isRunning():
            self.log("⚠️ Another task is running. Please wait...")
            return

        # Collect selected tasks
        selected_tasks = []
        for task_info in self.task_checkboxes:
            if task_info['checkbox'].isChecked():
                selected_tasks.append(task_info)

        if not selected_tasks:
            self.log("⚠️ No tasks selected. Please select at least one task.")
            return

        # Reset stop flag
        self.stop_requested = False

        # Update button states
        self.run_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)

        # Sort by original index to maintain order
        selected_tasks.sort(key=lambda x: x['index'])

        # Build actions dictionary based on selected tasks
        actions = {
            "handle_cookies": False,
            "create": False,
            "start": False,
            "pw": False,
            "import": False,
            "close": False,
            "delete": False,
        }

        # Map task names to actions
        for task_info in selected_tasks:
            name = task_info['name']
            if name == "Update Cookies to Excel":
                actions["handle_cookies"] = True
            elif name == "Create Profiles":
                actions["create"] = True
            elif name == "Import Cookies":
                actions["import"] = True
                actions["start"] = True
            elif name == "Start Profiles":
                actions["start"] = True
            elif name == "Start Profiles (with CDP)":
                actions["start"] = True
                actions["pw"] = True
            elif name == "Close Profiles":
                actions["close"] = True
            elif name == "Delete Profiles":
                actions["delete"] = True

        # Log selected tasks
        task_names = [f"{task['icon']} {task['name']}" for task in selected_tasks]
        self.log(f"\n{'='*70}")
        self.log(f"📋 Running tasks ({len(selected_tasks)}):")
        for i, name in enumerate(task_names, 1):
            self.log(f"  {i}. {name}")
        self.log(f"{'='*70}\n")

        # Run tasks with CLI logic
        self.run_tasks_with_actions(actions)

    def stop_tasks(self):
        """Stop currently running tasks"""
        if not self.current_worker or not self.current_worker.isRunning():
            self.log("⚠️ No tasks are currently running.")
            return

        self.log("\n🛑 Stopping tasks... Please wait...")
        self.stop_requested = True

        # Request worker to stop
        if self.current_worker:
            self.current_worker.stop()
            # Wait for worker to finish (with timeout)
            if self.current_worker.wait(5000):  # 5 seconds timeout
                self.log("✅ Tasks stopped successfully")
            else:
                self.log("⚠️ Force terminating tasks...")
                self.current_worker.terminate()
                self.current_worker.wait()

        # Reset button states
        self.run_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

    def run_tasks_with_actions(self, actions):
        """Run tasks using CLI logic with ThreadPoolExecutor"""
        def run():
            from concurrent.futures import ThreadPoolExecutor, as_completed
            import threading

            # Check if stop was requested
            if self.stop_requested or (self.current_worker and self.current_worker.is_stopped()):
                safe_print("⏹️ Task execution stopped by user")
                return

            # Initialize semaphore for start limit
            config.start_sem = threading.Semaphore(config.START_LIMIT)

            # Clear previous state
            with config.started_lock:
                config.started_debug_addrs.clear()

            with config.pw_jobs_lock:
                config.pw_jobs.clear()

            # Read excel data
            rows = read_excel()

            # Calculate total profiles for dynamic grid
            total_profiles = len(rows) if (actions["start"] or actions["import"] or actions["pw"]) else None

            # Execute with thread pool (same as CLI)
            with ThreadPoolExecutor(max_workers=config.THREADS) as ex:
                futures = [
                    ex.submit(process_row, *row[:3], i+1, actions, total_profiles)
                    for i, row in enumerate(rows)
                ]
                for future in as_completed(futures):
                    # Check stop request periodically
                    if self.stop_requested or (self.current_worker and self.current_worker.is_stopped()):
                        safe_print("⏹️ Stopping task execution...")
                        # Cancel remaining futures
                        for f in futures:
                            f.cancel()
                        ex.shutdown(wait=False)
                        return

            # Check before running playwright jobs
            if self.stop_requested or (self.current_worker and self.current_worker.is_stopped()):
                safe_print("⏹️ Task execution stopped before Playwright jobs")
                return

            # Run playwright jobs if needed
            if actions["import"] or actions["pw"]:
                with config.pw_jobs_lock:
                    jobs = config.pw_jobs.copy()

                if not jobs:
                    safe_print("⚠️ No playwright jobs collected.")
                else:
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    try:
                        loop.run_until_complete(run_all_playwright(jobs, actions))
                    except Exception as e:
                        if not self.stop_requested:
                            safe_print(f"❌ Playwright error: {e}")
                    finally:
                        loop.close()

            if not self.stop_requested and not (self.current_worker and self.current_worker.is_stopped()):
                safe_print("✅ ALL DONE")
            else:
                safe_print("⏹️ Tasks stopped by user")

        # Run in worker thread
        self.current_worker = TaskWorker("Automation Tasks", run)
        self.current_worker.log_signal.connect(self.log)
        self.current_worker.finished_signal.connect(self.on_tasks_finished)
        self.current_worker.start()

    @Slot(bool, str)
    def on_tasks_finished(self, success, message):
        """Handle completion of all tasks"""
        # Reset button states
        self.run_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

        if success:
            self.log("\n✅ All tasks completed successfully!")
        else:
            if "stopped" in message.lower():
                self.log(f"\n⏹️ {message}")
            else:
                self.log(f"\n❌ Tasks failed: {message}")

    def save_log(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Log",
            f"gpm_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            "Text Files (*.txt)"
        )
        if file_path:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(self.terminal.toPlainText())
            self.log(f"💾 Log saved: {file_path}")


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    window = GPMMainWindow()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
