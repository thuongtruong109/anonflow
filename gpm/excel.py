import os
from openpyxl import load_workbook

from config import EXCEL_PATH
from utils import list_filepaths

def update_excel_column_a_with_cookie_files(
    excel_path: str,
    cookies_folder: str,
    sheet_name: str | None = None,
    start_row: int = 2,
):
    cookie_paths = list_filepaths(cookies_folder)

    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    for row in range(start_row, ws.max_row + 1):
        ws.cell(row=row, column=1).value = None

    row = start_row
    for fpath in cookie_paths:
        ws.cell(row=row, column=1).value = fpath
        row += 1

def update_excel_column_c_with_profile_id(
    excel_path: str,
    profile_name: str,
    profile_id: str,
    sheet_name: str | None = None,
    start_row: int = 2,
):
    """Update column C (ID) for a specific profile name"""
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Find the row with matching profile name in column A
    for row in range(start_row, ws.max_row + 1):
        cookie_path = ws.cell(row=row, column=1).value
        if cookie_path:
            # Extract profile name from cookie path
            from utils import detect_username_from_cookie_filename
            current_profile_name = detect_username_from_cookie_filename(str(cookie_path))
            if current_profile_name == profile_name:
                # Update column C with the profile ID
                ws.cell(row=row, column=3).value = profile_id
                wb.save(excel_path)
                return True

    return False

def count_proxy_rows(
    excel_path: str,
    sheet_name: str | None = None,
    start_row: int = 2,
    proxy_col: int = 2,
) -> int:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    cnt = 0
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, proxy_col).value
        if v is None:
            continue
        if str(v).strip() == "":
            continue
        cnt += 1
    return cnt

def read_excel():
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = []

    for r in range(2, ws.max_row + 1):
        cookie_path = ws.cell(r, 1).value
        proxy = ws.cell(r, 2).value

        if proxy is None or str(proxy).strip() == "":
            continue

        cookie_path_str = str(cookie_path).strip() if cookie_path else ""
        profile_name = os.path.basename(cookie_path_str) if cookie_path_str else f"Profile {r-1}"

        rows.append((profile_name, cookie_path_str, proxy))

    return rows
